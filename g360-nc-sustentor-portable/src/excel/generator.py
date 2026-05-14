import os
import re
from typing import Dict, List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.formatting.rule import DataBarRule
from datetime import datetime
from src.core.processor import ProcessedItem
from src.core.utils import format_id_name, format_doc_id


class G360Styles:
    """
    Centraliza la identidad visual de los reportes G360.
    Define colores, fuentes y bordes compartidos entre plantillas y reportes finales.
    """
    def __init__(self):
        self.side = Side(style='thin', color="000000")
        self.border = Border(left=self.side, right=self.side, top=self.side, bottom=self.side)
        self.header_fill = PatternFill(start_color="0B1220", end_color="0B1220", fill_type="solid")
        self.header_font = Font(color="FFFFFF", bold=True)
        self.critical_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        self.total_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        self.alert_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        self.warning_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        self.info_fill = PatternFill(start_color="C9DAF8", end_color="C9DAF8", fill_type="solid")
        self.zebra_fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
        self.alert_font = Font(color="FFFFFF", bold=True)
        self.warning_font = Font(color="9C5700", bold=True)
        self.info_font = Font(color="003366", bold=True)
        self.center_align = Alignment(horizontal='center', vertical='center')
        self.left_align = Alignment(horizontal='left', vertical='center')
        self.wrap_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)


class ExcelGenerator:
    """
    Encargado de la transformación de ProcessedItems a archivos Excel (.xlsx).
    Implementa lógica de formato dinámico, fórmulas de Excel y auto-ajuste de columnas.
    """
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        assert self.ws is not None
        self.ws.title = "Sustento NC"
        self.styles = G360Styles()

        self.fmt_num = '#,##0.00'
        self.fmt_num_4 = '#,##0.0000'
        self.fmt_pct = '0.00%'

    def _limpiar(self, texto):
        """Asegura que el texto sea grabable en Excel (elimina caracteres no imprimibles)."""
        if texto is None: return ""
        return "".join(c for c in str(texto) if c.isprintable()).strip()

    def _format_doc_from_string(self, doc_str: str) -> str:
        """Helper para parsear strings tipo F201-1234567 y usar format_doc_id."""
        if '-' not in doc_str: return doc_str
        parts = doc_str.split('-', 1)
        prefix = parts[0] # F201
        nro = parts[1]    # 1234567
        tpo = prefix[0] if prefix else ""
        serie = prefix[1:] if len(prefix) > 1 else ""
        return format_doc_id(tpo, serie, nro)

    def _escribir_encabezado_y_totales(self, cliente: str, cliente_id: str = "", motivo: str = "", fila_fin_datos: int = 0, factura_ref: str = ""):
        """
        Construye la sección superior del reporte. 
        Utiliza referencias de celdas ($fila_fin_datos) para crear fórmulas de SUMA
        que abarquen exactamente el rango de datos procesados.
        """
        
        # Fila 1: Fecha
        self.ws.cell(row=1, column=1, value="FECHA:").font = Font(bold=True)
        self.ws.cell(row=1, column=2, value=datetime.now().strftime("%d/%m/%Y"))

        # Fila 2: Cliente con ID (formato: ID - NOMBRE)
        cliente_display = format_id_name(cliente_id, cliente).upper()
        c_cliente = self.ws.cell(row=2, column=1, value=cliente_display)
        c_cliente.font = Font(bold=True, size=14)
        self.ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)

        # Fila 3: Motivo
        self.ws.cell(row=3, column=1, value="MOTIVO:").font = Font(bold=True)
        self.ws.cell(row=3, column=2, value=self._limpiar(motivo))

        # Nota aclaratoria sobre IGV (Fila 4)
        c_nota = self.ws.cell(row=4, column=1, value="* Los cálculos de descuento y subtotales por ítem no incluyen IGV. El impuesto se calcula al finalizar el reporte.")
        c_nota.font = Font(italic=True, size=9, color="666666")
        self.ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=8)

        # Cuadro de Totales Superiores (Filas 1-3, Columnas J-K)
        # Los datos empiezan en la fila 7, por lo que la suma es de K7 a K...
        f_sub = f"=SUM(K7:K{max(8, fila_fin_datos)})"
        f_igv = f"=ROUND(K1*0.18, 2)"  # K1 es el Subtotal
        f_tot = f"=ROUND(K1+K2, 2)"    # K1 + K2 es el Total con IGV

        labels = [("Subtotal (Sin IGV):", f_sub), ("IGV (18.00%):", f_igv), ("TOTAL NC FINAL:", f_tot), ("FACTURA REF:", factura_ref)]
        
        for i, (lab, form) in enumerate(labels, 1):
            # Etiqueta (Columna J)
            c_l = self.ws.cell(row=i, column=10, value=lab)
            c_l.font = Font(bold=True)
            c_l.fill = self.styles.total_fill
            c_l.border = self.styles.border
            
            # Valor (Columna K)
            c_v = self.ws.cell(row=i, column=11, value=self._limpiar(form))
            c_v.number_format = self.fmt_num
            c_v.border = self.styles.border
            c_v.fill = self.styles.total_fill
            if "TOTAL" in lab:
                c_v.font = Font(bold=True, size=12)
            if "FACTURA" in lab:
                c_v.font = Font(bold=True, color="0000FF")
        
    def _escribir_cabeceras(self, fila: int):
        """Define los nombres de las columnas de la tabla de datos y aplica estilo G360."""
        cols = [
            "N°", "SKU", "SKU - ARTÍCULO", "LÍNEA", "CANT. SUSTENTAR", "P.U. (SIN IGV)",
            "TOT. FACT. (NETO)", "DESC. (%)", "DESC. UNIT. (NETO)", "PRECIO NETO",
            "SUBTOTAL (SIN IGV)", "FACTURAS", "ALERTA"
        ]
        for i, texto in enumerate(cols, 1):
            celda = self.ws.cell(row=fila, column=i, value=texto)
            celda.fill = self.styles.header_fill
            celda.font = self.styles.header_font
            celda.alignment = self.styles.center_align
            celda.border = self.styles.border

    def _escribir_fila(self, fila: int, item: ProcessedItem):
        """
        Escribe una fila de datos. Inserta fórmulas vivas (ROUND, SUM) en lugar de valores estáticos
        para permitir que el usuario realice ajustes manuales en el Excel si es necesario.
        Aplica lógica de colores (Semáforo de alertas) según el estado del ítem.
        """
        # Zebra Striping
        bg_fill = self.styles.zebra_fill if fila % 2 == 0 else None

        # Calcular índice (N°) para esta fila
        idx = fila - 7 + 1

        # Col 1: N°
        c_idx = self.ws.cell(row=fila, column=1, value=idx)
        c_idx.border = self.styles.border
        c_idx.alignment = self.styles.center_align

        # Col 2: SKU (ID Puro)
        c_sku_id = self.ws.cell(row=fila, column=2, value=str(item.ID_ARTICULO))
        c_sku_id.border = self.styles.border
        c_sku_id.alignment = self.styles.center_align

        # Col 3: SKU - ARTICULO
        sku_display = format_id_name(item.ID_ARTICULO, item.NOM_ARTICULO)
        c_sku_art = self.ws.cell(row=fila, column=3, value=sku_display)
        c_sku_art.border = self.styles.border

        # Col 4: LINEA (ID - NOMBRE)
        linea_display = format_id_name(item.ID_LINEA, item.NOM_LINEA)
        c_linea = self.ws.cell(row=fila, column=4, value=linea_display)
        c_linea.border = self.styles.border

        # Col 5: CANT. SUSTENTAR
        c_cant = self.ws.cell(row=fila, column=5, value=item.CANTIDAD_REAL_ENCONTRADA)
        c_cant.border = self.styles.border
        c_cant.alignment = self.styles.center_align

        # Col 6: P.U. (SIN IGV)
        c_pu = self.ws.cell(row=fila, column=6, value=float(item.PRECIO_UNITARIO))
        c_pu.border = self.styles.border
        c_pu.number_format = self.fmt_num

        # Col 7: TOT. FACT. (NETO) - Formula: CANT. SUSTENTAR * P.U.
        c_tf = self.ws.cell(row=fila, column=7, value=f"=ROUND(E{fila}*F{fila}, 2)")
        c_tf.border = self.styles.border
        c_tf.number_format = self.fmt_num

        # Col 8: DESC. (%)
        c_perc = self.ws.cell(row=fila, column=8, value=float(item.PORCENTAJE_APLICADO))
        c_perc.border = self.styles.border
        c_perc.number_format = self.fmt_pct
        c_perc.alignment = self.styles.center_align

        # Col 9: DESC. UNIT. (NETO) - Formula: P.U. * DESC. (%)
        c_du = self.ws.cell(row=fila, column=9, value=f"=ROUND(F{fila}*H{fila}, 4)")
        c_du.border = self.styles.border
        c_du.fill = self.styles.critical_fill
        c_du.number_format = self.fmt_num_4

        # Col 10: PRECIO NETO - Formula: P.U. - DESC. UNIT.
        c_neto = self.ws.cell(row=fila, column=10, value=f"=F{fila}-I{fila}")
        c_neto.border = self.styles.border
        c_neto.number_format = self.fmt_num_4

        # Col 11: SUBTOTAL (SIN IGV) - Formula: CANT. SUSTENTAR * DESC. UNIT.
        c_sub = self.ws.cell(row=fila, column=11, value=f"=ROUND(E{fila}*I{fila}, 2)")
        c_sub.border = self.styles.border
        c_sub.number_format = self.fmt_num
        
        formatted_docs = [self._format_doc_from_string(d) for d in item.DOCUMENTOS]
        c_docs = self.ws.cell(row=fila, column=12, value=self._limpiar("; ".join(formatted_docs)))
        c_docs.border = self.styles.border
        c_docs.alignment = self.styles.wrap_alignment
        
        status = self._limpiar(item.STATUS)
        # Aplicar Zebra Striping a toda la fila si corresponde
        if bg_fill: # Columnas A hasta M (13 columns)
            for col_idx in range(1, 14):
                if col_idx != 9: # No sobreescribir el color crítico de la columna I (DESC. UNIT)
                    self.ws.cell(row=fila, column=col_idx).fill = bg_fill
                if col_idx == 9:
                    self.ws.cell(row=fila, column=col_idx).fill = self.styles.critical_fill

        c_alert = self.ws.cell(row=fila, column=13, value=status)
        c_alert.border = self.styles.border
        c_alert.alignment = self.styles.wrap_alignment
        
        # Lógica de colores por tipo de alerta
        if any(x in status.upper() for x in ["ERROR", "ALERTA", "PENDIENTE", "FALTAN"]):
            c_alert.fill = self.styles.alert_fill
            c_alert.font = self.styles.alert_font
        elif "VARIABLE" in status.upper():
            c_alert.fill = self.styles.warning_fill
            c_alert.font = self.styles.warning_font
        elif "INFO" in status.upper():
            c_alert.fill = self.styles.info_fill
            c_alert.font = self.styles.info_font

    def generar_reporte(self, ruta_salida: str, cliente: str, cliente_id: str = "", motivo: str = "",
                     items_procesados=None, documentos_unicos=None, rango_fechas=None, 
                     sheet_name=None, factura_referencia=""):
        """
        Genera un reporte de Notas de Crédito en formato Excel.
        
        Args:
            ruta_salida (str): Ruta completa donde se guardará el archivo Excel.
            cliente (str): Nombre del cliente para el encabezado del reporte.
            cliente_id (str): ID del cliente para mostrar en formato "ID - NOMBRE".
            motivo (str): Motivo de la Nota de Crédito.
            items_procesados (List[ProcessedItem]): Lista de ítems ya procesados por NCProcessor.
            documentos_unicos (List[str]): Lista de documentos únicos utilizados en el sustento.
            rango_fechas (Tuple[Optional[pd.Timestamp], Optional[pd.Timestamp]]): Rango de fechas del historial.
            sheet_name (Optional[str]): Nombre opcional para la hoja de Excel.
        """
        # ✅ SOLUCION 100% COMPATIBLE CON TODAS LAS VERSIONES DE OPENPYXL
        os.makedirs(os.path.dirname(os.path.abspath(ruta_salida)), exist_ok=True)

        # 1. Calcular fila final real
        fila_inicio_datos = 7
        fila_fin_datos = fila_inicio_datos + len(items_procesados) - 1

        # Asignar nombre a la hoja si se proporciona (limpiando caracteres prohibidos en Excel)        
        if sheet_name:
            clean_name = re.sub(r'[\\/*?:\[\]]', "", str(sheet_name))[:31]
            if clean_name:
                self.ws.title = clean_name

        # 2. Escribir Encabezado y Totales Superiores
        self._escribir_encabezado_y_totales(cliente, cliente_id, motivo, fila_fin_datos, factura_referencia)

        self.ws.freeze_panes = "D7" # Congelar ID y Nombre, y filas de encabezado

        # 3. Cabeceras de Tabla (Fila 6)
        fila_cab = 6
        self._escribir_cabeceras(fila_cab)
        
        # 4. Datos (Fila 7 en adelante)
        f_act = 7
        for it in items_procesados:
            self._escribir_fila(f_act, it)
            f_act += 1
        
        # 5. Footer
        f_foot = f_act + 1
        self.ws.merge_cells(start_row=f_foot, start_column=1, end_row=f_foot, end_column=12)
        txt_docs = f"Documentos únicos procesados: {', '.join([self._limpiar(d) for d in documentos_unicos])}"
        c_f = self.ws.cell(row=f_foot, column=1, value=txt_docs)
        c_f.font = Font(italic=True, color="555555")

        # 6. Auto-ajuste de anchos optimizado (muestreo de las primeras 100 filas)
        for col in self.ws.columns:
            max_length = 0
            column = col[0].column_letter
            # Solo verificamos las cabeceras y las primeras 100 filas para rendimiento
            for i, cell in enumerate(col):
                if i > 100: break 
                try:
                    if cell.value:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except: pass
            adjusted_width = (max_length + 3)
            self.ws.column_dimensions[column].width = min(adjusted_width, 50) # Cap at 50
            
        self.wb.save(str(ruta_salida))

    def generar_plantilla_vacia(self, ruta_salida):
        """
        Genera la plantilla oficial de Requerimientos lista para usar,
        con formato, ejemplos, validaciones e instrucciones.
        """
        # ✅ SOLUCION 100% COMPATIBLE CON TODAS LAS VERSIONES DE OPENPYXL
        os.makedirs(os.path.dirname(os.path.abspath(ruta_salida)), exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "REQUERIMIENTOS"

        note_fill = PatternFill(start_color="FFFBE6", end_color="FFFBE6", fill_type="solid")

        # Cabeceras oficiales
        columnas = [
            ("CODIGO", "Código del Artículo / SKU"),
            ("NOM_ARTICULO", "Nombre del Artículo (opcional)"),
            ("CANTIDAD_NC", "Cantidad de unidades a procesar"),
            ("PORCENTAJE_DESC", "Descuento a aplicar (%)")
        ]

        # Forzar formatos de celda
        for r in range(2, 501):
            ws.cell(row=r, column=1).number_format = '@'       # CODIGO
            ws.cell(row=r, column=3).number_format = '0'       # CANTIDAD
            ws.cell(row=r, column=4).number_format = '0.00%'   # PORCENTAJE

        # Escribir cabeceras con estilo
        for col, (nombre, descripcion) in enumerate(columnas, 1):
            celda = ws.cell(row=1, column=col, value=nombre)
            celda.fill = self.styles.header_fill
            celda.font = self.styles.header_font
            celda.alignment = self.styles.center_align
            celda.comment = f"\n{descripcion}\n"
            celda.border = self.styles.border

        # Ejemplos de uso (con los datos solicitados)
        ejemplos = [
            ["123456", "PRODUCTO EJEMPLO 1", 5, "10%"],
            ["789012", "PRODUCTO EJEMPLO 2", 12, 3.5],
            ["345678", "", 2, 0.05],
            ["02182", "PRODUCTO EJEMPLO 1", 423, 0.05],
            ["123456", "PRODUCTO EJEMPLO 2", 12, 0.10],
        ]

        for fila, datos in enumerate(ejemplos, 2):
            for col, valor in enumerate(datos, 1):
                celda = ws.cell(row=fila, column=col, value=valor)
                celda.fill = self.styles.zebra_fill
                celda.border = self.styles.border

        # Filas de instrucciones
        fila_nota = 6
        ws.merge_cells(start_row=fila_nota, start_column=1, end_row=fila_nota, end_column=4)
        celda_nota = ws.cell(row=fila_nota, column=1, value="📋 INSTRUCCIONES:")
        celda_nota.font = Font(bold=True, size=11)
        celda_nota.fill = note_fill

        instrucciones = [
            "1. Eliminar las filas de ejemplo (2,3,4) antes de cargar tus datos",
            "2. Solo las columnas CODIGO, CANTIDAD_NC y PORCENTAJE_DESC son obligatorias",
            "3. El descuento se acepta en formato: 10%, 10, 0.1, 10.5",
            "4. No dejar filas vacías entre registros",
            "5. No modificar el nombre ni orden de las columnas",
            "6. Guardar el archivo antes de importar al sistema"
        ]

        for i, texto in enumerate(instrucciones, 7):
            ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=4)
            celda = ws.cell(row=i, column=1, value=texto)
            celda.font = Font(size=10, color="444444")

        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 22
        ws.column_dimensions['D'].width = 25

        # Congelar primera fila
        ws.freeze_panes = "A2"

        # Agregar filtros automáticos
        ws.auto_filter.ref = "A1:D1"

        wb.save(str(ruta_salida))

    def generar_reporte_consolidado_excel(self, ruta_salida: str, titulo: str,
                                         datos_por_vendedor: Dict[str, List[Dict]], 
                                         agrupacion: str, rango_fechas: tuple,
                                         historial=None, filtros_aplicados: str = ""):
        """Genera Excel con múltiples hojas: RESUMEN + una por vendedor."""
        self.wb = Workbook()
        
        # Verificar si es Pareto por Cliente
        if agrupacion == "PARETO_CLIENTE":
            self._escribir_pareto_cliente(titulo, datos_por_vendedor, rango_fechas)
            
            # Generar anexo de sucursales
            from src.reports.consolidated import ReporteConsolidado
            anexo = ReporteConsolidado.generar_pareto_sucursales(
                historial=historial,
                clientes_filtro=None,
                vendedores_filtro=None,
                lineas_filtro=None
            )
            if anexo:
                self._escribir_pareto_sucursales(titulo, anexo, rango_fechas, datos_por_vendedor)
            
            self.wb.save(str(ruta_salida))
            return
        
        # Construir diccionario de nombres de vendedores desde el historial
        vendedores_nombres = {}
        if historial is not None and 'ID_VENDEDOR' in historial.columns and 'NOM_VENDEDOR' in historial.columns:
            for _, row in historial[['ID_VENDEDOR', 'NOM_VENDEDOR']].drop_duplicates(subset='ID_VENDEDOR').iterrows():
                vendedores_nombres[row['ID_VENDEDOR']] = str(row['NOM_VENDEDOR'])
        
        summary_sheet = self.wb.create_sheet("RESUMEN")
        self._escribir_resumen_consolidado(summary_sheet, titulo, datos_por_vendedor, rango_fechas, vendedores_nombres, filtros_aplicados)
        
        for vendedor_id, datos in datos_por_vendedor.items():
            sheet_name = f"V_{vendedor_id}"[:31]
            ws = self.wb.create_sheet(sheet_name)
            self.ws = ws
            vendedor_nombre = vendedores_nombres.get(vendedor_id, "")
            self._escribir_encabezado_consolidado(ws, titulo, vendedor_id, rango_fechas, vendedor_nombre)
            
            dato_ejemplo = datos[0] if datos else {}
            
            if 'FACTURA' in dato_ejemplo:
                cabeceras = ['N°', 'FACTURA', 'FECHA', 'CLIENTE', 'LÍNEA', 'SKU', 'CANTIDAD', 'PRECIO', 'MONTO']
            elif 'TENDENCIA' in dato_ejemplo or any(k.endswith('-MONTO') for k in dato_ejemplo):
                # Reporte Comparativo / Evolución mensual
                cabeceras = ['N°', 'SKU', 'LÍNEA', 'CLIENTE']
                # Meses dinámicos
            elif 'PERIODO' in dato_ejemplo and 'TIPO_PERIODO' in dato_ejemplo:
                # Por PERIODO
                cabeceras = ['N°', 'PERIODO', 'FECHA', 'SKU', 'LÍNEA', 'CLIENTES', 'CANTIDAD', 'MONTO', 'FACTURAS', 'PRECIOS']
            elif 'ID_CLIENTE' in dato_ejemplo and 'CLIENTE' in dato_ejemplo and 'ID_LINEA' in dato_ejemplo:
                # Por CLIENTE
                cabeceras = ['N°', 'CLIENTE', 'LÍNEA', 'SKU', 'CANTIDAD', 'MONTO', 'FECHA ULT.', 'FACTURAS', 'PRECIOS']
            elif 'ID_LINEA' in dato_ejemplo and 'LÍNEA' in dato_ejemplo and 'SKU' in dato_ejemplo:
                # Por LÍNEA
                cabeceras = ['N°', 'LÍNEA', 'SKU', 'CLIENTES', 'CANTIDAD', 'MONTO', 'FECHA ULT.', 'FACTURAS', 'PRECIOS']
            elif 'SKU' in dato_ejemplo and 'CLIENTES' in dato_ejemplo:
                # Por SKU
                cabeceras = ['N°', 'SKU', 'LÍNEA', 'CLIENTES', 'CANTIDAD', 'MONTO', 'FECHA ULT.', 'FACTURAS', 'PRECIOS']
            else:
                # Fallback
                cabeceras = ['N°', 'CLIENTE', 'LÍNEA', 'SKU', 'CANTIDAD', 'MONTO', 'FECHA ULT.', 'FACTURAS', 'PRECIOS']

            # Escribir cabeceras con estilo (Fila 5)
            # Manejo especial para meses comparativos (merged headers en fila 4)
            if 'TENDENCIA' in dato_ejemplo or any(k.endswith('-MONTO') for k in dato_ejemplo):
                meses_cols = sorted(set(k.rsplit('-', 1)[0] for k in dato_ejemplo.keys() if '-' in k and any(s in k for s in ['-MONTO', '-CANT', '-PRECIO'])))
                for idx_mes, mes in enumerate(meses_cols):
                    start_col = 5 + (idx_mes * 3)
                    ws.cell(row=4, column=start_col, value=mes.upper()).font = Font(bold=True, size=11)
                    ws.cell(row=4, column=start_col).alignment = self.styles.center_align
                    ws.merge_cells(start_row=4, start_column=start_col, end_row=4, end_column=start_col + 2)
                    cabeceras.extend([f'CANTIDAD', f'MONTO', f'PRECIO'])
                if 'DIF_SOLES' in dato_ejemplo:
                    cabeceras.extend(['DIF_SOLES', 'DIF_PCT'])
                if 'TENDENCIA' in dato_ejemplo:
                    cabeceras.extend(['TENDENCIA'])

            for i, cab in enumerate(cabeceras, 1):
                celda = ws.cell(row=5, column=i, value=cab)
                celda.fill = self.styles.header_fill
                celda.font = self.styles.header_font
                celda.border = self.styles.border
                celda.alignment = self.styles.center_align

            for fila_idx, dato in enumerate(datos, 6):
                col = 1
                idx = fila_idx - 5

                if 'FACTURA' in dato:
                    # Reporte POR FACTURA
                    ws.cell(row=fila_idx, column=col, value=idx).border = self.styles.border
                    ws.cell(row=fila_idx, column=col).alignment = self.styles.center_align
                    col += 1
                    ws.cell(row=fila_idx, column=col, value=dato.get('FACTURA', '')).border = self.styles.border; col+=1
                    ws.cell(row=fila_idx, column=col, value=dato.get('FECHA', '')).border = self.styles.border; col+=1
                    ws.cell(row=fila_idx, column=col, value=dato.get('CLIENTE', '')).border = self.styles.border; col+=1
                    ws.cell(row=fila_idx, column=col, value=dato.get('LÍNEA', '')).border = self.styles.border; col+=1
                    ws.cell(row=fila_idx, column=col, value=dato.get('SKU', '')).border = self.styles.border; col+=1
                    c_cant = ws.cell(row=fila_idx, column=col, value=dato.get('CANTIDAD', 0))
                    c_cant.border = self.styles.border
                    c_cant.number_format = '#,##0'
                    col += 1
                    c_precio = ws.cell(row=fila_idx, column=col, value=dato.get('PRECIO', 0))
                    c_precio.border = self.styles.border
                    c_precio.number_format = '#,##0.00'
                    col += 1
                    c_monto = ws.cell(row=fila_idx, column=col, value=dato.get('MONTO', 0))
                    c_monto.border = self.styles.border
                    c_monto.number_format = '#,##0.00'
                    col += 1
                ws.cell(row=fila_idx, column=col, value=dato.get('FECHA_ULT', '')).border = self.styles.border; col += 1
                elif 'TENDENCIA' in dato or any(k.endswith('-MONTO') for k in dato.keys()):
                    # Reporte COMPARATIVO / EVOLUCIÓN MENSUAL - columnas dinámicas
                    ws.cell(row=fila_idx, column=col, value=idx).border = self.styles.border
                    ws.cell(row=fila_idx, column=col).alignment = self.styles.center_align
                    col += 1
                    ws.cell(row=fila_idx, column=col, value=dato.get('SKU', '')).border = self.styles.border; col+=1
                    ws.cell(row=fila_idx, column=col, value=dato.get('LÍNEA', '')).border = self.styles.border; col+=1
                    ws.cell(row=fila_idx, column=col, value=dato.get('CLIENTE', '')).border = self.styles.border; col+=1
                    # Escribir columnas de meses
                    meses_cols = set()
                    for key in dato.keys():
                        if '-MONTO' in key or '-CANT' in key or '-PRECIO' in key:
                            meses_cols.add(key.rsplit('-', 1)[0])
                    for mes in sorted(meses_cols):
                        ws.cell(row=fila_idx, column=col, value=dato.get(f'{mes}-CANT', 0)).border = self.styles.border; col+=1
                        ws.cell(row=fila_idx, column=col, value=dato.get(f'{mes}-MONTO', 0)).border = self.styles.border; col+=1
                        ws.cell(row=fila_idx, column=col, value=dato.get(f'{mes}-PRECIO', 0)).border = self.styles.border; col+=1
                    if 'DIF_SOLES' in dato:
                        ws.cell(row=fila_idx, column=col, value=dato.get('DIF_SOLES', 0)).border = self.styles.border; col+=1
                        ws.cell(row=fila_idx, column=col, value=dato.get('DIF_PCT', 0)).border = self.styles.border; col+=1
                    if 'TENDENCIA' in dato:
                        ws.cell(row=fila_idx, column=col, value=dato.get('TENDENCIA', '')).border = self.styles.border; col+=1
                elif 'PERIODO' in dato:
                    # Reporte POR PERIODO
                    ws.cell(row=fila_idx, column=col, value=idx).border = self.styles.border
                    ws.cell(row=fila_idx, column=col).alignment = self.styles.center_align
                    col += 1
                    ws.cell(row=fila_idx, column=col, value=dato.get('PERIODO', '')).border = self.styles.border; col+=1
                    fecha_val = dato.get('FECHA', dato.get('FECHA_MIN', ''))
                    ws.cell(row=fila_idx, column=col, value=fecha_val[:10] if fecha_val else '').border = self.styles.border; col+=1
                    ws.cell(row=fila_idx, column=col, value=dato.get('SKU', '')).border = self.styles.border; col+=1
                    ws.cell(row=fila_idx, column=col, value=dato.get('LÍNEA', '')).border = self.styles.border; col+=1
                    ws.cell(row=fila_idx, column=col, value=dato.get('CLIENTES', '')).border = self.styles.border; col+=1
                    c_cant = ws.cell(row=fila_idx, column=col, value=dato.get('CANTIDAD', 0))
                    c_cant.border = self.styles.border
                    c_cant.number_format = '#,##0'
                    col += 1
                    c_monto = ws.cell(row=fila_idx, column=col, value=dato.get('MONTO', 0))
                    c_monto.border = self.styles.border
                    c_monto.number_format = '#,##0.00'
                    col += 1
                ws.cell(row=fila_idx, column=col, value=dato.get('FECHA_ULT', '')).border = self.styles.border; col += 1
                    ws.cell(row=fila_idx, column=col, value=dato.get('FACTURAS', '')).border = self.styles.border; col+=1
                    ws.cell(row=fila_idx, column=col, value=dato.get('PRECIOS', '')).border = self.styles.border; col+=1
                elif 'ID_CLIENTE' in dato and 'CLIENTE' in dato:
                    # Reporte POR CLIENTE: N°, CLIENTE, LÍNEA, SKU, CANTIDAD, MONTO, FACTURAS, PRECIOS
                    ws.cell(row=fila_idx, column=col, value=idx).border = self.styles.border
                    ws.cell(row=fila_idx, column=col).alignment = self.styles.center_align
                    col += 1
                    ws.cell(row=fila_idx, column=col, value=dato.get('CLIENTE', '')).border = self.styles.border; col+=1
                    ws.cell(row=fila_idx, column=col, value=dato.get('LÍNEA', '')).border = self.styles.border; col+=1
                    ws.cell(row=fila_idx, column=col, value=dato.get('SKU', '')).border = self.styles.border; col+=1
                    c_cant = ws.cell(row=fila_idx, column=col, value=dato.get('CANTIDAD', 0))
                    c_cant.border = self.styles.border
                    c_cant.number_format = '#,##0'
                    col += 1
                    c_monto = ws.cell(row=fila_idx, column=col, value=dato.get('MONTO', 0))
                    c_monto.border = self.styles.border
                    c_monto.number_format = '#,##0.00'
                    col += 1
                ws.cell(row=fila_idx, column=col, value=dato.get('FECHA_ULT', '')).border = self.styles.border; col += 1
                    ws.cell(row=fila_idx, column=col, value=dato.get('FACTURAS', '')).border = self.styles.border; col+=1
                    ws.cell(row=fila_idx, column=col, value=dato.get('PRECIOS', '')).border = self.styles.border; col+=1
                elif 'ID_LINEA' in dato:
                    # Reporte POR LÍNEA: N°, LÍNEA, SKU, CLIENTES, CANTIDAD, MONTO, FACTURAS, PRECIOS
                    ws.cell(row=fila_idx, column=col, value=idx).border = self.styles.border
                    ws.cell(row=fila_idx, column=col).alignment = self.styles.center_align
                    col += 1
                    ws.cell(row=fila_idx, column=col, value=dato.get('LÍNEA', '')).border = self.styles.border; col+=1
                    ws.cell(row=fila_idx, column=col, value=dato.get('SKU', '')).border = self.styles.border; col+=1
                    ws.cell(row=fila_idx, column=col, value=dato.get('CLIENTES', '')).border = self.styles.border; col+=1
                    c_cant = ws.cell(row=fila_idx, column=col, value=dato.get('CANTIDAD', 0))
                    c_cant.border = self.styles.border
                    c_cant.number_format = '#,##0'
                    col += 1
                    c_monto = ws.cell(row=fila_idx, column=col, value=dato.get('MONTO', 0))
                    c_monto.border = self.styles.border
                    c_monto.number_format = '#,##0.00'
                    col += 1
                ws.cell(row=fila_idx, column=col, value=dato.get('FECHA_ULT', '')).border = self.styles.border; col += 1
                    ws.cell(row=fila_idx, column=col, value=dato.get('FACTURAS', '')).border = self.styles.border; col+=1
                    ws.cell(row=fila_idx, column=col, value=dato.get('PRECIOS', '')).border = self.styles.border; col+=1
                elif 'SKU' in dato and 'CLIENTES' in dato:
                    # Reporte POR SKU: N°, SKU, LÍNEA, CLIENTES, CANTIDAD, MONTO, FACTURAS, PRECIOS
                    ws.cell(row=fila_idx, column=col, value=idx).border = self.styles.border
                    ws.cell(row=fila_idx, column=col).alignment = self.styles.center_align
                    col += 1
                    ws.cell(row=fila_idx, column=col, value=dato.get('SKU', '')).border = self.styles.border; col+=1
                    ws.cell(row=fila_idx, column=col, value=dato.get('LÍNEA', '')).border = self.styles.border; col+=1
                    ws.cell(row=fila_idx, column=col, value=dato.get('CLIENTES', '')).border = self.styles.border; col+=1
                    c_cant = ws.cell(row=fila_idx, column=col, value=dato.get('CANTIDAD', 0))
                    c_cant.border = self.styles.border
                    c_cant.number_format = '#,##0'
                    col += 1
                    c_monto = ws.cell(row=fila_idx, column=col, value=dato.get('MONTO', 0))
                    c_monto.border = self.styles.border
                    c_monto.number_format = '#,##0.00'
                    col += 1
                ws.cell(row=fila_idx, column=col, value=dato.get('FECHA_ULT', '')).border = self.styles.border; col += 1
                    ws.cell(row=fila_idx, column=col, value=dato.get('FACTURAS', '')).border = self.styles.border; col+=1
                    ws.cell(row=fila_idx, column=col, value=dato.get('PRECIOS', '')).border = self.styles.border; col+=1
                elif 'PRODUCTO' in dato:
                    # Reporte POR PRODUCTO / ARTÍCULO
                    ws.cell(row=fila_idx, column=col, value=idx).border = self.styles.border
                    ws.cell(row=fila_idx, column=col).alignment = self.styles.center_align
                    col += 1
                    ws.cell(row=fila_idx, column=col, value=dato.get('PRODUCTO', '')).border = self.styles.border; col+=1
                    ws.cell(row=fila_idx, column=col, value=dato.get('SKU', '')).border = self.styles.border; col+=1
                    ws.cell(row=fila_idx, column=col, value=dato.get('LÍNEA', '')).border = self.styles.border; col+=1
                    ws.cell(row=fila_idx, column=col, value=dato.get('CLIENTES', '')).border = self.styles.border; col+=1
                    c_cant = ws.cell(row=fila_idx, column=col, value=dato.get('CANTIDAD', 0))
                    c_cant.border = self.styles.border
                    c_cant.number_format = '#,##0'
                    col += 1
                    c_monto = ws.cell(row=fila_idx, column=col, value=dato.get('MONTO', 0))
                    c_monto.border = self.styles.border
                    c_monto.number_format = '#,##0.00'
                    col += 1
                ws.cell(row=fila_idx, column=col, value=dato.get('FECHA_ULT', '')).border = self.styles.border; col += 1
                    ws.cell(row=fila_idx, column=col, value=dato.get('FACTURAS', '')).border = self.styles.border; col+=1
                    ws.cell(row=fila_idx, column=col, value=dato.get('PRECIOS', '')).border = self.styles.border; col+=1
                else:
                    # Fallback genérico: N°, CLIENTE, LÍNEA, SKU, CANTIDAD, MONTO, FACTURAS, PRECIOS
                    ws.cell(row=fila_idx, column=col, value=idx).border = self.styles.border
                    ws.cell(row=fila_idx, column=col).alignment = self.styles.center_align
                    col += 1
                    ws.cell(row=fila_idx, column=col, value=dato.get('CLIENTE', '')).border = self.styles.border; col+=1
                    ws.cell(row=fila_idx, column=col, value=dato.get('LÍNEA', '')).border = self.styles.border; col+=1
                    ws.cell(row=fila_idx, column=col, value=dato.get('SKU', '')).border = self.styles.border; col+=1
                    c_cant = ws.cell(row=fila_idx, column=col, value=dato.get('CANTIDAD', 0))
                    c_cant.border = self.styles.border
                    c_cant.number_format = '#,##0'
                    col += 1
                    c_monto = ws.cell(row=fila_idx, column=col, value=dato.get('MONTO', 0))
                    c_monto.border = self.styles.border
                    c_monto.number_format = '#,##0.00'
                    col += 1
                ws.cell(row=fila_idx, column=col, value=dato.get('FECHA_ULT', '')).border = self.styles.border; col += 1
                    ws.cell(row=fila_idx, column=col, value=dato.get('FACTURAS', '')).border = self.styles.border; col+=1
                    ws.cell(row=fila_idx, column=col, value=dato.get('PRECIOS', '')).border = self.styles.border; col+=1
            
            self._auto_adjust_columns(ws)
            self._aplicar_formato_condicional_monto(ws, cabeceras, fila_idx)
        
        if len(self.wb.sheetnames) > 1 and "Sheet" in self.wb.sheetnames:
            del self.wb["Sheet"]
        
        self.wb.save(str(ruta_salida))

    def _escribir_resumen_consolidado(self, ws, titulo, datos_por_vendedor, rango_fechas, vendedores_nombres=None, filtros_aplicados: str = ""):
        ws.cell(row=1, column=1, value=titulo).font = Font(bold=True, size=14)
        ws.cell(row=2, column=1, value=f"Fecha: {datetime.now().strftime('%d/%m/%Y')}")
        if rango_fechas and rango_fechas[0] and rango_fechas[1]:
            ws.cell(row=3, column=1, value=f"Rango: {rango_fechas[0].strftime('%d/%m/%Y')} -> {rango_fechas[1].strftime('%d/%m/%Y')}")
        
        if filtros_aplicados:
            ws.cell(row=4, column=1, value=f"Filtros: {filtros_aplicados}").font = Font(italic=True, color="666666")

        if vendedores_nombres is None:
            vendedores_nombres = {}
        cabeceras = ['VENDEDOR', 'ITEMS', 'MONTO_TOTAL']
        for i, cab in enumerate(cabeceras, 1):
            celda = ws.cell(row=5, column=i, value=cab)
            celda.fill = self.styles.header_fill
            celda.font = self.styles.header_font
        fila = 6
        for vendedor_id, datos in datos_por_vendedor.items():
            total = sum(d.get('MONTO', 0) for d in datos)
            nombre = vendedores_nombres.get(vendedor_id, "")
            ws.cell(row=fila, column=1, value=format_id_name(vendedor_id, nombre))
            ws.cell(row=fila, column=2, value=len(datos))
            ws.cell(row=fila, column=3, value=total).number_format = '#,##0.00'
            fila += 1

    def _escribir_encabezado_consolidado(self, ws, titulo, vendedor_id, rango_fechas, vendedor_nombre=""):
        ws.cell(row=1, column=1, value=titulo).font = Font(bold=True, size=14)
        ws.cell(row=2, column=1, value=f"Vendedor: {format_id_name(vendedor_id, vendedor_nombre)}")
        ws.cell(row=3, column=1, value=f"Fecha: {datetime.now().strftime('%d/%m/%Y')}")
        if rango_fechas and rango_fechas[0] and rango_fechas[1]:
            ws.cell(row=4, column=1, value=f"Periodo: {rango_fechas[0].strftime('%d/%m/%Y')} -> {rango_fechas[1].strftime('%d/%m/%Y')}")

    def _aplicar_formato_condicional_monto(self, ws, cabeceras, max_row):
        """Aplica barras de datos a las columnas de MONTO."""
        from openpyxl.utils import get_column_letter
        for i, cab in enumerate(cabeceras, 1):
            if 'MONTO' in cab.upper():
                col_letter = get_column_letter(i)
                rule = DataBarRule(start_type='min', end_type='max', color="00D084", showValue=True)
                ws.conditional_formatting.add(f"{col_letter}6:{col_letter}{max_row}", rule)

    def _auto_adjust_columns(self, ws):
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for i, cell in enumerate(col):
                if i > 100: break
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            ws.column_dimensions[column].width = min(max_length + 3, 50)
    
    def _escribir_pareto_cliente(self, titulo: str, datos_por_vendedor: Dict, rango_fechas: tuple):
        '''Escribe reporte Pareto por Cliente (una hoja por vendedor).'''
        if 'Sheet' in self.wb.sheetnames:
            del self.wb['Sheet']
        
        for vendedor_id, datos in datos_por_vendedor.items():
            if not datos:
                continue
            
            sheet_name = f'V_{vendedor_id[:28]}'[:31]
            ws = self.wb.create_sheet(sheet_name)
            self.ws = ws
            
            ws.cell(row=1, column=1, value=f'{titulo} - Pareto por Cliente').font = Font(bold=True, size=14)
            ws.cell(row=2, column=1, value=f'Vendedor: {datos[0].get("VENDEDOR", vendedor_id)}')
            ws.cell(row=3, column=1, value=f'Fecha: {datetime.now().strftime("%d/%m/%Y")}')
            
            if rango_fechas and rango_fechas[0] and rango_fechas[1]:
                ws.cell(row=4, column=1, value=f'Periodo: {rango_fechas[0].strftime("%d/%m/%Y")} → {rango_fechas[1].strftime("%d/%m/%Y")}')
            
            meses = set()
            for dato in datos:
                for key in dato.keys():
                    if key.endswith('-CANT') or key.endswith('-MONTO'):
                        mes = key.rsplit('-', 1)[0]
                        meses.add(mes)
            meses_ordenados = sorted(meses)
            
            cabeceras = ['N°', 'CLIENTE', 'CANTIDAD', 'MONTO', '% IND', '% ACUM', 'CATEGORIA']
            if datos and 'SUCURSAL' in datos[0]:
                cabeceras.insert(2, 'SUCURSAL')
            if datos and 'FACTURAS' in datos[0]:
                cabeceras.insert(len(cabeceras), 'FACTURAS')
            
            for mes in meses_ordenados:
                cabeceras.extend([f'{mes}-CANT', f'{mes}-MONTO'])
            
            if datos and 'TENDENCIA' in datos[0]:
                cabeceras.extend(['TENDENCIA'])
            
            for i, cab in enumerate(cabeceras, 1):
                celda = ws.cell(row=6, column=i, value=cab)
                celda.fill = self.styles.header_fill
                celda.font = self.styles.header_font
                celda.border = self.styles.border
                celda.alignment = self.styles.center_align
            
            for idx, dato in enumerate(datos, 1):
                fila_idx = idx + 6
                col = 1
                
                ws.cell(row=fila_idx, column=col, value=idx).border = self.styles.border
                ws.cell(row=fila_idx, column=col).alignment = self.styles.center_align
                col += 1
                
                ws.cell(row=fila_idx, column=col, value=dato.get('CLIENTE', '')).border = self.styles.border
                col += 1
                
                if 'SUCURSAL' in dato:
                    ws.cell(row=fila_idx, column=col, value=dato.get('SUCURSAL', '')).border = self.styles.border
                    col += 1
                
                c_cant = ws.cell(row=fila_idx, column=col, value=dato.get('CANTIDAD', 0))
                c_cant.border = self.styles.border
                c_cant.number_format = '#,##0'
                col += 1
                
                c_monto = ws.cell(row=fila_idx, column=col, value=dato.get('MONTO', 0))
                c_monto.border = self.styles.border
                c_monto.number_format = '#,##0.00'
                col += 1
                ws.cell(row=fila_idx, column=col, value=dato.get('FECHA_ULT', '')).border = self.styles.border; col += 1
                
                c_pct_ind = ws.cell(row=fila_idx, column=col, value=dato.get('PCT_INDIVIDUAL', 0) / 100)
                c_pct_ind.border = self.styles.border
                c_pct_ind.number_format = '0.00%'
                col += 1
                
                c_pct_acum = ws.cell(row=fila_idx, column=col, value=dato.get('PCT_ACUMULADO', 0) / 100)
                c_pct_acum.border = self.styles.border
                c_pct_acum.number_format = '0.00%'
                col += 1
                
                ws.cell(row=fila_idx, column=col, value=dato.get('CATEGORIA', '')).border = self.styles.border
                col += 1
                
                if 'FACTURAS' in dato:
                    ws.cell(row=fila_idx, column=col, value=dato.get('FACTURAS', '')).border = self.styles.border
                    col += 1
                
                for mes in meses_ordenados:
                    ws.cell(row=fila_idx, column=col, value=dato.get(f'{mes}-CANT', 0)).border = self.styles.border
                    col += 1
                    ws.cell(row=fila_idx, column=col, value=dato.get(f'{mes}-MONTO', 0)).border = self.styles.border
                    col += 1
                
                if 'TENDENCIA' in dato:
                    ws.cell(row=fila_idx, column=col, value=dato.get('TENDENCIA', '')).border = self.styles.border
                    col += 1
            
            ws.column_dimensions['A'].width = 25
            self._auto_adjust_columns(ws)
    
    def _escribir_pareto_sucursales(self, titulo: str, datos_por_vendedor: Dict, rango_fechas: tuple, datos_pareto: Dict = None):
        '''Escribe reporte Anexo por Cliente + Sucursal (mantiene secuencia del Pareto).'''
        for vendedor_id, datos in datos_por_vendedor.items():
            if not datos:
                continue
            
            orden_clientes = []
            if datos_pareto and vendedor_id in datos_pareto:
                for idx, d in enumerate(datos_pareto[vendedor_id]):
                    cliente = d.get('CLIENTE', '')
                    if cliente not in orden_clientes:
                        orden_clientes.append(cliente)
            
            clientes_contados = {}
            for dato in datos:
                cliente = dato.get('CLIENTE', '')
                clientes_contados[cliente] = clientes_contados.get(cliente, 0) + 1
            
            datos_filtrados = [d for d in datos if clientes_contados.get(d.get('CLIENTE', ''), 0) >= 2]
            
            if orden_clientes:
                datos_filtrados.sort(key=lambda x: (
                    orden_clientes.index(x['CLIENTE']) if x['CLIENTE'] in orden_clientes else 999,
                    -x['MONTO_TOTAL']
                ))
            else:
                datos_filtrados.sort(key=lambda x: (x['CLIENTE'], -x['MONTO_TOTAL']))
            
            if not datos_filtrados:
                continue
            
            sheet_name = f'Anexo_{vendedor_id[:25]}'[:31]
            if sheet_name in self.wb.sheetnames:
                del self.wb[sheet_name]
            ws = self.wb.create_sheet(sheet_name)
            self.ws = ws
            
            ws.cell(row=1, column=1, value=titulo + ' - Anexo Sucursales').font = Font(bold=True, size=14)
            ws.cell(row=2, column=1, value=f'Vendedor: {datos[0].get("VENDEDOR", vendedor_id)}')
            ws.cell(row=3, column=1, value=f'Fecha: {datetime.now().strftime("%d/%m/%Y")}')
            
            if rango_fechas and rango_fechas[0] and rango_fechas[1]:
                ws.cell(row=4, column=1, value=f'Periodo: {rango_fechas[0].strftime("%d/%m/%Y")} → {rango_fechas[1].strftime("%d/%m/%Y")}')
            
            meses = set()
            for dato in datos_filtrados:
                for key in dato.keys():
                    if key.endswith('-CANT') or key.endswith('-MONTO'):
                        meses.add(key.rsplit('-', 1)[0])
            meses_ordenados = sorted(meses)
            
            cabeceras = ['N°', 'CLIENTE', 'SUCURSAL', 'FACTURAS', 'CANTIDAD', 'MONTO', '% IND', '% ACUM', 'CATEGORIA']
            for mes in meses_ordenados:
                cabeceras.extend([f'{mes}-CANT', f'{mes}-MONTO'])
            
            for i, cab in enumerate(cabeceras, 1):
                celda = ws.cell(row=6, column=i, value=cab)
                celda.fill = self.styles.header_fill
                celda.font = self.styles.header_font
                celda.border = self.styles.border
                celda.alignment = self.styles.center_align
            
            fila_idx = 7
            idx = 0
            cliente_anterior = None
            for dato in datos_filtrados:
                cliente = dato.get('CLIENTE', '')
                if cliente != cliente_anterior:
                    idx += 1
                    cliente_anterior = cliente
                
                col = 1
                ws.cell(row=fila_idx, column=col, value=idx).border = self.styles.border
                ws.cell(row=fila_idx, column=col).alignment = self.styles.center_align
                col += 1
                
                ws.cell(row=fila_idx, column=col, value=cliente).border = self.styles.border
                col += 1
                
                ws.cell(row=fila_idx, column=col, value=dato.get('SUCURSAL', '')).border = self.styles.border
                col += 1
                
                c_fact = ws.cell(row=fila_idx, column=col, value=dato.get('FACTURAS', ''))
                c_fact.border = self.styles.border
                c_fact.alignment = Alignment(wrap_text=True)
                col += 1
                
                c_cant = ws.cell(row=fila_idx, column=col, value=dato.get('CANTIDAD', 0))
                c_cant.border = self.styles.border
                c_cant.number_format = '#,##0'
                col += 1
                
                c_monto = ws.cell(row=fila_idx, column=col, value=dato.get('MONTO', 0))
                c_monto.border = self.styles.border
                c_monto.number_format = '#,##0.00'
                col += 1
                ws.cell(row=fila_idx, column=col, value=dato.get('FECHA_ULT', '')).border = self.styles.border; col += 1
                
                c_pct = ws.cell(row=fila_idx, column=col, value=dato.get('PCT_INDIVIDUAL', 0) / 100)
                c_pct.border = self.styles.border
                c_pct.number_format = '0.00%'
                col += 1
                
                ws.cell(row=fila_idx, column=col, value=dato.get('PCT_ACUMULADO', 0) / 100).border = self.styles.border
                col += 1
                
                ws.cell(row=fila_idx, column=col, value=dato.get('CATEGORIA', '')).border = self.styles.border
                col += 1
                
                for mes in meses_ordenados:
                    ws.cell(row=fila_idx, column=col, value=dato.get(f'{mes}-CANT', 0)).border = self.styles.border
                    col += 1
                    ws.cell(row=fila_idx, column=col, value=dato.get(f'{mes}-MONTO', 0)).border = self.styles.border
                    col += 1
                
                fila_idx += 1
            
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 18
            ws.column_dimensions['C'].width = 40
            self._auto_adjust_columns(ws)
