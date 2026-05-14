import os
import re
import pandas as pd
from typing import Dict, List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.formatting.rule import DataBarRule
from datetime import datetime
from src.core.processor import ProcessedItem
from src.core.utils import format_id_name, format_doc_id, IGV_PERCENT
from src.core.data_dictionary import DataDictionary


class G360Styles:
    """
    Centraliza la identidad visual de los reportes G360.
    Define colores, fuentes y bordes compartidos entre plantillas y reportes finales.
    """
    def __init__(self):
        self.side = Side(style='thin', color="000000")
        self.border = Border(left=self.side, right=self.side, top=self.side, bottom=self.side)
        self.header_fill = PatternFill(start_color="0D2B4E", end_color="0D2B4E", fill_type="solid")
        self.header_font = Font(color="FFFFFF", bold=True, size=10)
        self.critical_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        self.total_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
        self.alert_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        self.warning_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        self.info_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        self.zebra_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        self.alert_font = Font(color="9C0006", bold=True, size=10)
        self.warning_font = Font(color="9C5700", bold=True, size=10)
        self.info_font = Font(color="003366", bold=True, size=10)
        self.center_align = Alignment(horizontal='center', vertical='center')
        self.left_align = Alignment(horizontal='left', vertical='center')
        self.right_align = Alignment(horizontal='right', vertical='center')
        self.wrap_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Estilos adicionales para reutilizar
        self.title_font = Font(bold=True, size=14)
        self.title_font_green = Font(bold=True, size=16, color="00D084")
        self.bold_font = Font(bold=True)
        self.italic_gray_font = Font(italic=True, color="666666")
        self.trend_up_font = Font(color="008000", bold=True)
        self.trend_down_font = Font(color="FF0000", bold=True)
        self.note_fill = PatternFill(start_color="FFFBE6", end_color="FFFBE6", fill_type="solid")
        
        # Estilos para alarmas HHI
        self.hhi_high_fill = PatternFill(start_color="9C0006", end_color="9C0006", fill_type="solid")
        self.hhi_high_font = Font(bold=True, size=12, color="FFFFFF")
        self.hhi_mod_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        self.hhi_mod_font = Font(bold=True, size=12, color="9C5700")
        self.hhi_low_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        self.hhi_low_font = Font(bold=True, size=12, color="006100")
        
        # Estilos para subtotales y fórmulas
        self.subtotal_font = Font(bold=True, size=11)
        self.subtotal_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
        self.kpi_value_fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
        self.kpi_header_fill = PatternFill(start_color="0D2B4E", end_color="0D2B4E", fill_type="solid")
        self.kpi_label_font = Font(color="FFFFFF", bold=True, size=11)
        self.kpi_value_font = Font(bold=True, size=12, color="003366")
        
        # Estilos para notas de crédito
        self.dev_font = Font(color="9C0006")
        self.dev_sub_font = Font(bold=True, size=11, color="9C0006")
        self.dev_label_font = Font(bold=True, size=10, color="9C0006")
        
        # Estilos para vulnerabilidad
        self.vuln_high_font = Font(bold=True, color="9C0006")
        self.vuln_mod_font = Font(bold=True, color="9C5700")
        
        # Estilo para alertas críticas (texto blanco sobre fondo rojo)
        self.critical_white_font = Font(bold=True, color="FFFFFF")


class ExcelGenerator:
    """
    Encargado de la transformación de ProcessedItems a archivos Excel (.xlsx).
    Implementa lógica de formato dinámico, fórmulas de Excel y auto-ajuste de columnas.
    """
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        assert self.ws is not None
        self.ws.title = "Sustento NC"
        self.styles = G360Styles()

        self.fmt_num = '#,##0.00'
        self.fmt_num_4 = '#,##0.0000'
        self.fmt_pct = '0.00%'

    def _limpiar(self, texto):
        """Asegura que el texto sea grabable en Excel (elimina caracteres no imprimibles)."""
        if texto is None: return ""
        return "".join(c for c in str(texto) if c.isprintable()).strip()

    def _format_doc_from_string(self, doc_str: str) -> str:
        """Helper para parsear strings tipo F201-1234567 y usar format_doc_id."""
        if '-' not in doc_str: return doc_str
        parts = doc_str.split('-', 1)
        prefix = parts[0] # F201
        nro = parts[1]    # 1234567
        tpo = prefix[0] if prefix else ""
        serie = prefix[1:] if len(prefix) > 1 else ""
        return format_doc_id(tpo, serie, nro)

    def _escribir_encabezado_y_totales(self, cliente: str, cliente_id: str = "", motivo: str = "", fila_fin_datos: int = 0, factura_ref: str = ""):
        """
        Construye la sección superior del reporte. 
        Utiliza referencias de celdas ($fila_fin_datos) para crear fórmulas de SUMA
        que abarquen exactamente el rango de datos procesados.
        """
        
        # Fila 1: Fecha
        self.ws.cell(row=1, column=1, value="FECHA:").font = self.styles.bold_font
        self.ws.cell(row=1, column=2, value=datetime.now().strftime("%d/%m/%Y"))

        # Fila 2: Cliente con ID (formato: ID - NOMBRE)
        cliente_display = format_id_name(cliente_id, cliente, field_name='CLIENTE').upper()
        c_cliente = self.ws.cell(row=2, column=1, value=cliente_display)
        c_cliente.font = Font(bold=True, size=14)
        self.ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)

        # Fila 3: Motivo
        self.ws.cell(row=3, column=1, value="MOTIVO:").font = self.styles.bold_font
        self.ws.cell(row=3, column=2, value=self._limpiar(motivo))

        # Nota aclaratoria sobre IGV (Fila 4)
        c_nota = self.ws.cell(row=4, column=1, value="* Los cálculos de descuento y subtotales por ítem no incluyen IGV. El impuesto se calcula al finalizar el reporte.")
        c_nota.font = Font(italic=True, size=9, color="666666")
        self.ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=8)

        # Cuadro de Totales Superiores (Filas 1-3, Columnas J-K)
        # Los datos empiezan en la fila 7, por lo que la suma es de K7 a K...
        f_sub = f"=SUM(K7:K{max(8, fila_fin_datos)})"
        f_igv = f"=ROUND(K1*{IGV_PERCENT}, 2)"  # K1 es el Subtotal
        f_tot = f"=ROUND(K1+K2, 2)"    # K1 + K2 es el Total con IGV
        factura_display = factura_ref if factura_ref else "SIN REFERENCIA"
        labels = [("Subtotal (Sin IGV):", f_sub), ("IGV (18.00%):", f_igv), ("TOTAL NC FINAL:", f_tot), ("FACTURA REF:", factura_display)]
        
        for i, (lab, form) in enumerate(labels, 1):
            # Etiqueta (Columna J)
            c_l = self.ws.cell(row=i, column=10, value=lab)
            c_l.font = self.styles.bold_font
            c_l.fill = self.styles.total_fill
            c_l.border = self.styles.border
            
            # Valor (Columna K)
            c_v = self.ws.cell(row=i, column=11, value=self._limpiar(form))
            if i == 3:  # TOTAL NC FINAL
                c_v.font = Font(bold=True, size=12)
            elif i == 4:  # FACTURA REF
                c_v.font = Font(bold=True, color="0000FF")
            c_v.border = self.styles.border
            c_v.fill = self.styles.total_fill
            if "TOTAL" in lab:
                c_v.font = Font(bold=True, size=12)
            if "FACTURA" in lab:
                c_v.font = Font(bold=True, color="0000FF")
        
    def _escribir_cabeceras(self, fila: int):
        """Define los nombres de las columnas de la tabla de datos y aplica estilo G360."""
        cols = [
            "N°", "SKU", "SKU - ARTÍCULO", "LÍNEA", "CANT. SUSTENTAR", "P.U. (SIN IGV)",
            "TOT. FACT. (NETO)", "DESC. (%)", "DESC. UNIT. (NETO)", "PRECIO NETO",
            "SUBTOTAL (SIN IGV)", "FACTURAS", "ALERTA"
        ]
        for i, texto in enumerate(cols, 1):
            celda = self.ws.cell(row=fila, column=i, value=texto)
            celda.fill = self.styles.header_fill
            celda.font = self.styles.header_font
            celda.alignment = self.styles.center_align
            celda.border = self.styles.border

    def _escribir_fila(self, fila: int, item: ProcessedItem, sku_count: int = 1):
        """
        Escribe una fila de datos con fórmulas vivas.
        Nueva disposición: SKU en Col B, SKU - ARTÍCULO en Col C, desplazando el resto a la derecha.
        Escribe una fila de datos. Inserta fórmulas vivas (ROUND, SUM) en lugar de valores estáticos
        para permitir que el usuario realice ajustes manuales en el Excel si es necesario.
        Aplica lógica de colores (Semáforo de alertas) según el estado del ítem.
        Si sku_count > 1, agrega alerta de SKU duplicado.
        """
        # Zebra Striping
        bg_fill = self.styles.zebra_fill if fila % 2 == 0 else None
        
        # Calcular índice (N°) para esta fila (Fila 7 = N°1, Fila 8 = N°2, etc.)
        idx = fila - 7 + 1
        
        # Col A (1): N°
        c_idx = self.ws.cell(row=fila, column=1, value=idx)
        c_idx.border = self.styles.border
        c_idx.alignment = self.styles.center_align

        # Col B (2): SKU (ID Puro)
        c_sku_id = self.ws.cell(row=fila, column=2, value=str(item.ID_ARTICULO))
        c_sku_id.border = self.styles.border
        c_sku_id.alignment = self.styles.center_align
        
        # Col C (3): SKU - ARTICULO
        sku_display = format_id_name(item.ID_ARTICULO, item.NOM_ARTICULO, field_name='SKU')
        c_sku_art = self.ws.cell(row=fila, column=3, value=sku_display)
        c_sku_art.border = self.styles.border
        
        # Col D (4): LINEA (ID - NOMBRE)
        linea_display = format_id_name(item.ID_LINEA, item.NOM_LINEA, field_name='LÍNEA')
        c_linea = self.ws.cell(row=fila, column=4, value=linea_display)
        c_linea.border = self.styles.border
        
        # Col E (5): CANT. SUSTENTAR
        c_cant = self.ws.cell(row=fila, column=5, value=item.CANTIDAD_REAL_ENCONTRADA)
        c_cant.border = self.styles.border
        c_cant.alignment = self.styles.center_align
        
        # Col F (6): P.U. (SIN IGV)
        c_pu = self.ws.cell(row=fila, column=6, value=float(item.PRECIO_UNITARIO))
        c_pu.border = self.styles.border
        c_pu.number_format = self.fmt_num
        
        # Col G (7): TOT. FACT. (NETO) - Formula: CANT. * P.U.
        c_tf = self.ws.cell(row=fila, column=7, value=f"=ROUND(E{fila}*F{fila}, 2)")
        c_tf.border = self.styles.border
        c_tf.number_format = self.fmt_num
        
        # Col H (8): DESC. (%)
        c_perc = self.ws.cell(row=fila, column=8, value=float(item.PORCENTAJE_APLICADO))
        c_perc.border = self.styles.border
        c_perc.number_format = self.fmt_pct
        c_perc.alignment = self.styles.center_align
        
        # Col I (9): DESC. UNIT. (NETO) - Formula: P.U. * DESC. (%)
        c_du = self.ws.cell(row=fila, column=9, value=f"=ROUND(F{fila}*H{fila}, 4)")
        c_du.border = self.styles.border
        c_du.fill = self.styles.critical_fill
        c_du.number_format = self.fmt_num_4
        
        # Col J (10): PRECIO NETO - Formula: P.U. - DESC. UNIT.
        c_neto = self.ws.cell(row=fila, column=10, value=f"=F{fila}-I{fila}")
        c_neto.border = self.styles.border
        c_neto.number_format = self.fmt_num_4
        
        # Col K (11): SUBTOTAL (SIN IGV) - Formula: CANT. * DESC. UNIT.
        c_sub = self.ws.cell(row=fila, column=11, value=f"=ROUND(E{fila}*I{fila}, 2)")
        c_sub.border = self.styles.border
        c_sub.number_format = self.fmt_num
        
        # Col L (12): FACTURAS
        formatted_docs = [self._format_doc_from_string(d) for d in item.DOCUMENTOS]
        c_docs = self.ws.cell(row=fila, column=12, value=self._limpiar("; ".join(formatted_docs)))
        c_docs.border = self.styles.border
        c_docs.alignment = self.styles.wrap_alignment
        
        # Col M (13): ALERTA
        status = self._limpiar(item.STATUS)
        if sku_count > 1:
            status = f"⚠️ SKU DUPLICADO ({sku_count}x): {status}" if status else f"⚠️ SKU DUPLICADO ({sku_count}x)"
        
        # Zebra Striping para columnas B-M
        if bg_fill:
            for col_idx in range(2, 14):
                if col_idx != 9:
                    if bg_fill: # Defensive check
                        self.ws.cell(row=fila, column=col_idx).fill = bg_fill
                if col_idx == 9:
                    self.ws.cell(row=fila, column=col_idx).fill = self.styles.critical_fill
        
        c_alert = self.ws.cell(row=fila, column=13, value=status)
        c_alert.border = self.styles.border
        c_alert.alignment = self.styles.wrap_alignment
        
        # Lógica de colores por tipo de alerta
        if any(x in status.upper() for x in ["ERROR", "ALERTA", "FALTAN", "DUPLICADO"]):
            c_alert.fill = self.styles.alert_fill
            c_alert.font = self.styles.alert_font
        elif "VARIABLE" in status.upper() or "PENDIENTE" in status.upper():
            c_alert.fill = self.styles.warning_fill
            c_alert.font = self.styles.warning_font
        elif "INFO" in status.upper():
            c_alert.fill = self.styles.info_fill
            c_alert.font = self.styles.info_font

    def generar_reporte(self, ruta_salida: str, cliente: str, cliente_id: str = "", motivo: str = "",
                     items_procesados=None, documentos_unicos=None, rango_fechas=None, 
                     sheet_name=None, factura_referencia=""):
        """
        Genera un reporte de Notas de Crédito en formato Excel.
        
        Args:
            ruta_salida (str): Ruta completa donde se guardará el archivo Excel.
            cliente (str): Nombre del cliente para el encabezado del reporte.
            cliente_id (str): ID del cliente para mostrar en formato "ID - NOMBRE".
            motivo (str): Motivo de la Nota de Crédito.
            items_procesados (List[ProcessedItem]): Lista de ítems ya procesados por NCProcessor.
            documentos_unicos (List[str]): Lista de documentos únicos utilizados en el sustento.
            rango_fechas (Tuple[Optional[pd.Timestamp], Optional[pd.Timestamp]]): Rango de fechas del historial.
            sheet_name (Optional[str]): Nombre opcional para la hoja de Excel.
        """
        # ✅ SOLUCION 100% COMPATIBLE CON TODAS LAS VERSIONES DE OPENPYXL
        os.makedirs(os.path.dirname(os.path.abspath(ruta_salida)), exist_ok=True)

        # 1. Calcular fila final real
        fila_inicio_datos = 7
        fila_fin_datos = fila_inicio_datos + len(items_procesados) - 1

        # Asignar nombre a la hoja si se proporciona (limpiando caracteres prohibidos en Excel)        
        if sheet_name:
            clean_name = re.sub(r'[\\/*?:\[\]]', "", str(sheet_name))[:31]
            if clean_name:
                self.ws.title = clean_name

        # 2. Escribir Encabezado y Totales Superiores
        self._escribir_encabezado_y_totales(cliente, cliente_id, motivo, fila_fin_datos, factura_referencia)

        self.ws.freeze_panes = "D7" # Congelar ID y Nombre, y filas de encabezado

        # 3. Cabeceras de Tabla (Fila 6)
        fila_cab = 6
        self._escribir_cabeceras(fila_cab)
        
        # 4. Detectar SKU duplicados
        sku_counts = {}
        for it in items_procesados:
            sku = it.ID_ARTICULO
            sku_counts[sku] = sku_counts.get(sku, 0) + 1
        
        # 5. Datos (Fila 7 en adelante)
        f_act = 7
        for it in items_procesados:
            self._escribir_fila(f_act, it, sku_counts.get(it.ID_ARTICULO, 1))
            f_act += 1
        
        # 5. Footer
        f_foot = f_act + 1
        self.ws.merge_cells(start_row=f_foot, start_column=1, end_row=f_foot, end_column=12)
        txt_docs = f"Documentos únicos procesados: {', '.join([self._limpiar(d) for d in documentos_unicos])}"
        c_f = self.ws.cell(row=f_foot, column=1, value=txt_docs)
        c_f.font = Font(italic=True, color="555555")

        # 6. Auto-ajuste de anchos
        self._auto_adjust_columns(self.ws)
            
        try:
            self.wb.save(str(ruta_salida))
        except PermissionError:
            raise PermissionError(
                f"No se pudo guardar el archivo. ¿Está abierto en otro programa?\n"
                f"Cierre el archivo e intente nuevamente.\nRuta: {ruta_salida}"
            )

    def generar_plantilla_vacia(self, ruta_salida):
        """
        Genera la plantilla oficial de Requerimientos lista para usar,
        con formato, ejemplos, validaciones e instrucciones.
        """
        # ✅ SOLUCION 100% COMPATIBLE CON TODAS LAS VERSIONES DE OPENPYXL
        os.makedirs(os.path.dirname(os.path.abspath(ruta_salida)), exist_ok=True)

        wb = Workbook()
        # Forzar modo de cálculo automático y recalcular al abrir
        wb.calculation.calcMode = 'auto'
        wb.calculation.fullCalcOnLoad = True
        
        ws = wb.active
        ws.title = "REQUERIMIENTOS"

        # Cabeceras oficiales
        columnas = [
            ("CODIGO", "Código del Artículo / SKU"),
            ("NOM_ARTICULO", "Nombre del Artículo (opcional)"),
            ("CANTIDAD_NC", "Cantidad de unidades a procesar"),
            ("PORCENTAJE_DESC", "Descuento a aplicar (%)")
        ]

        # Forzar formatos de celda
        for r in range(2, 501):
            ws.cell(row=r, column=1).number_format = '@'       # CODIGO
            ws.cell(row=r, column=3).number_format = '0'       # CANTIDAD
            ws.cell(row=r, column=4).number_format = '0.00%'   # PORCENTAJE

        # Escribir cabeceras con estilo
        for col, (nombre, descripcion) in enumerate(columnas, 1):
            celda = ws.cell(row=1, column=col, value=nombre)
            celda.fill = self.styles.header_fill
            celda.font = self.styles.header_font
            celda.alignment = self.styles.center_align
            celda.comment = f"\n{descripcion}\n"
            celda.border = self.styles.border

        # Ejemplos de uso (con los datos solicitados)
        ejemplos = [
            ["123456", "PRODUCTO EJEMPLO 1", 5, "10%"],
            ["789012", "PRODUCTO EJEMPLO 2", 12, 3.5],
            ["345678", "", 2, 0.05],
            ["02182", "PRODUCTO EJEMPLO 1", 423, 0.05],
            ["123456", "PRODUCTO EJEMPLO 2", 12, 0.10],
        ]

        for fila, datos in enumerate(ejemplos, 2):
            for col, valor in enumerate(datos, 1):
                celda = ws.cell(row=fila, column=col, value=valor)
                celda.fill = self.styles.zebra_fill
                celda.border = self.styles.border

        # Filas de instrucciones
        fila_nota = 6
        ws.merge_cells(start_row=fila_nota, start_column=1, end_row=fila_nota, end_column=4)
        celda_nota = ws.cell(row=fila_nota, column=1, value="📋 INSTRUCCIONES:")
        celda_nota.font = Font(bold=True, size=11)
        celda_nota.fill = self.styles.note_fill

        instrucciones = [
            "1. Eliminar las filas de ejemplo (2,3,4) antes de cargar tus datos",
            "2. Solo las columnas CODIGO, CANTIDAD_NC y PORCENTAJE_DESC son obligatorias",
            "3. El descuento se acepta en formato: 10%, 10, 0.1, 10.5",
            "4. No dejar filas vacías entre registros",
            "5. No modificar el nombre ni orden de las columnas",
            "6. Guardar el archivo antes de importar al sistema"
        ]

        for i, texto in enumerate(instrucciones, 7):
            ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=4)
            celda = ws.cell(row=i, column=1, value=texto)
            celda.font = Font(size=10, color="444444")

        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 22
        ws.column_dimensions['D'].width = 25

        # Congelar primera fila
        ws.freeze_panes = "A2"

        # Agregar filtros automáticos
        ws.auto_filter.ref = "A1:D1"

        try:
            wb.save(str(ruta_salida))
        except PermissionError:
            raise PermissionError(
                f"No se pudo guardar la plantilla. ¿Está abierto en otro programa?\n"
                f"Ruta: {ruta_salida}"
            )

    def generar_reporte_consolidado_excel(self, ruta_salida: str, titulo: str,
                                         datos_por_vendedor: Dict[str, List[Dict]], 
                                         agrupacion: str, rango_fechas: tuple,
                                         historial=None, filtros_aplicados: str = ""):
        """Genera Excel con múltiples hojas: RESUMEN + una por vendedor."""
        from openpyxl.utils import get_column_letter
        self.wb = Workbook()
        
        # Orquestación de Reporte Pareto (Multivista)
        if agrupacion == "PARETO_CLIENTE":
            # Usamos los datos ya procesados y filtrados que vienen de la UI
            datos_pareto = datos_por_vendedor 
            
            # Obtener líneas de forma segura
            lineas = datos_pareto.get('LINEAS', []) or []
            
            # 1. Hoja Resumen Global
            self._escribir_resumen_pareto_v2(datos_pareto, rango_fechas)
            
        # 2. Una hoja por Vendedor (INTEGRADA: Pareto + Líneas + Sucursales + Evolución)
            por_vendedor = datos_pareto.get('POR_VENDEDOR', {}) or {}
            from src.reports.consolidated import ReporteConsolidado
            try:
                datos_sucursales_all = ReporteConsolidado.generar_pareto_sucursales(historial=historial)
            except Exception:
                datos_sucursales_all = {}
            
            for v_display, items in por_vendedor.items():
                # Extraer ID del vendedor y primera palabra del nombre
                partes = v_display.split(' - ') if ' - ' in v_display else (v_display.split('-') if '-' in v_display else [v_display])
                id_vendedor = partes[0] if partes else v_display
                nom_vendedor = partes[1] if len(partes) > 1 else ""
                primer_palabra = nom_vendedor.split(' ')[0] if nom_vendedor else id_vendedor
                sheet_name = f"{id_vendedor}_{primer_palabra}"[:31]
                sheet_name = re.sub(r'[\\/*?:\[\]]', "", sheet_name)
                if sheet_name in self.wb.sheetnames:
                    sheet_name = f"{sheet_name}_copy"[:31]
                
                # Obtener datos de sucursales para este vendedor (si existe)
                datos_suc = datos_sucursales_all.get(id_vendedor, [])
                
                self._escribir_hoja_vendedor_integrada(
                    self.wb.create_sheet(sheet_name),
                    v_display, items or [], lineas,
                    historial, datos_suc, rango_fechas
                )

            try:
                if 'Sheet' in self.wb.sheetnames: del self.wb['Sheet']
                self.wb.save(str(ruta_salida))
            except PermissionError:
                raise PermissionError(f"Cierre el archivo {ruta_salida} antes de generar.")
            return
        
        # 1. Identificar nombres de vendedores para el resumen
        vendedores_nombres = {}
        # Obtener nombres desde el historial si está disponible
        if historial is not None and 'ID_VENDEDOR' in historial.columns and 'NOM_VENDEDOR' in historial.columns:
            vendedores_nombres = dict(zip(historial['ID_VENDEDOR'], historial['NOM_VENDEDOR']))
        else:
            for v_id, items in datos_por_vendedor.items():
                if items:
                    vendedores_nombres[v_id] = items[0].get('VENDEDOR', v_id)
        
        # 2. Hoja de Resumen (siempre la primera)
        ws_resumen = self.wb.active
        ws_resumen.title = "RESUMEN"
        self._escribir_resumen_consolidado(ws_resumen, titulo, datos_por_vendedor, rango_fechas, vendedores_nombres, filtros_aplicados)
        self._auto_adjust_columns(ws_resumen)
        
        # 3. Una hoja por vendedor (formato: ID_PRIMERA_PALABRA)
        for vendedor_id, items in datos_por_vendedor.items():
            if not items: continue
            
            v_nombre = vendedores_nombres.get(vendedor_id, vendedor_id)
            # Extraer primera palabra del nombre, limpiar tildes y caracteres especiales
            primer_palabra = v_nombre.split(' ')[0] if v_nombre else vendedor_id
            primer_palabra = re.sub(r'[\\/*?:\[\]]', "", primer_palabra)
            # Limpiar nombre para la hoja (max 31 chars, sin caracteres prohibidos)
            sheet_name = f"{vendedor_id}_{primer_palabra}"[:31]
            
            ws = self.wb.create_sheet(sheet_name)
            self._escribir_encabezado_consolidado(ws, titulo, vendedor_id, rango_fechas, v_nombre)
            
            # Detectar campos dinámicamente basándose en los datos
            dato_ejemplo = items[0] if items else {}
            cabeceras = self._detectar_cabeceras(dato_ejemplo, agrupacion)
            
            # Si la agrupación es por FACTURA, asegurar que tenga la columna FACTURA
            if agrupacion == "FACTURA" and "FACTURA" not in cabeceras:
                cabeceras.insert(0, "FACTURA")
            
            # Si tiene tendencia, agregar columnas de tendencia
            if 'TENDENCIA' in dato_ejemplo:
                if 'DIF_SOLES' not in cabeceras:
                    cabeceras.append('DIF_SOLES')
                if 'DIF_PCT' not in cabeceras:
                    cabeceras.append('DIF_PCT')
                if 'TENDENCIA' not in cabeceras:
                    cabeceras.append('TENDENCIA')
            
            # Escribir cabeceras
            fila_cab = 6
            for i, text in enumerate(cabeceras, 1):
                celda = ws.cell(row=fila_cab, column=i, value=text)
                celda.fill = self.styles.header_fill
                celda.font = self.styles.header_font
                celda.border = self.styles.border
                celda.alignment = self.styles.center_align
            
            # Datos: escribir según los campos detectados
            fila = 7
            for it in items:
                self._escribir_fila_consolidado(ws, fila, it, cabeceras)
                fila += 1
            
            # Formatos finales por hoja
            self._aplicar_formato_condicional_monto(ws, cabeceras, fila - 1)
            self._auto_adjust_columns(ws)
            ws.freeze_panes = "A7"
        # 4. Guardar archivo final
        try:
            self.wb.save(str(ruta_salida))
        except PermissionError:
            raise PermissionError(f"No se pudo guardar el archivo. ¿Está abierto?\nRuta: {ruta_salida}")

    def _escribir_resumen_pareto_v2(self, datos_pareto, rango_fechas):
        from openpyxl.formatting.rule import DataBarRule
        ws = self.wb.active
        ws.title = "RESUMEN"
        kpis = datos_pareto.get('RESUMEN', {}).get('KPIS', {})
        clientes = datos_pareto.get('CLIENTES', [])
        
        ws.cell(row=1, column=1, value="RESUMEN EJECUTIVO PARETO (80/20)").font = self.styles.title_font_green
        ws.cell(row=2, column=1, value=f"Análisis del historial: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
        
        kpi_list = [
            ("Facturación Total", kpis.get('MONTO_TOTAL', 0), "S/ #,##0.00"),
            ("Clientes Atendidos", kpis.get('TOTAL_CLIENTES', 0), "#,##0"),
            ("Clientes Vitales (80%)", kpis.get('CLIENTES_VITALES', 0), "#,##0"),
            ("Market Share Vitales", kpis.get('PCT_VITAL_MARKET_SHARE', 0) / 100, "0.00%"), # Nuevo KPI
            ("Líneas de Negocio", kpis.get('TOTAL_LINEAS', 0), "#,##0"),
            ("Total Documentos", kpis.get('TOTAL_FACTURAS', 0), "#,##0"),
        ]
        
        for i, (label, val, fmt) in enumerate(kpi_list, 4):
            c1 = ws.cell(row=i, column=1, value=label)
            c1.font = self.styles.kpi_label_font; c1.border = self.styles.border; c1.fill = self.styles.kpi_header_fill
            c2 = ws.cell(row=i, column=2, value=val)
            c2.font = self.styles.kpi_value_font; c2.number_format = fmt; c2.border = self.styles.border; c2.fill = self.styles.kpi_value_fill

        # Escribir Tabla Global de Clientes
        fila_h = len(kpi_list) + 5
        ws.cell(row=fila_h, column=1, value="TOP CLIENTES GLOBAL").font = self.styles.title_font
        
        fila_h += 1
        cabeceras = ['N°', 'CLIENTE', 'VENDEDOR', 'TOTAL SOLES', '% GLOB', 'TEND', 'CAT']
        for i, text in enumerate(cabeceras, 1):
            c = ws.cell(row=fila_h, column=i, value=text)
            c.fill = self.styles.header_fill; c.font = self.styles.header_font; c.border = self.styles.border; c.alignment = self.styles.center_align
        
        fila = fila_h + 1
        
        for idx, cli in enumerate(clientes):
            es_vital = 'VITAL' in str(cli.get('CATEGORIA', '')).upper()
            nro = idx + 1
            
            # Col A: N°
            c0 = ws.cell(row=fila, column=1, value=nro)
            c0.border = self.styles.border; c0.alignment = self.styles.center_align
            
            cli_str = cli.get('CLIENTE', f"{cli.get('ID_CLIENTE', '')} - {cli.get('NOM_CLIENTE', '')}".strip(" -"))
            c1 = ws.cell(row=fila, column=2, value=cli_str)
            c1.border = self.styles.border
            
            ven_str = cli.get('VENDEDOR', f"{cli.get('ID_VENDEDOR', '')} - {cli.get('NOM_VENDEDOR', '')}".strip(" -"))
            c2 = ws.cell(row=fila, column=3, value=ven_str)
            c2.border = self.styles.border
            
            c3 = ws.cell(row=fila, column=4, value=cli.get('MONTO_TOTAL'))
            c3.number_format = '#,##0.00'; c3.border = self.styles.border
            
            c4 = ws.cell(row=fila, column=5, value=cli.get('PCT_GLOBAL', 0)/100)
            c4.number_format = '0.00%'; c4.border = self.styles.border
            
            c5 = ws.cell(row=fila, column=6, value=cli.get('TENDENCIA', '➡️'))
            c5.alignment = self.styles.center_align; c5.border = self.styles.border
            if '🔺' in str(c5.value): c5.font = self.styles.trend_up_font
            elif '🔻' in str(c5.value): c5.font = self.styles.trend_down_font
            
            c6 = ws.cell(row=fila, column=7, value=cli.get('CATEGORIA'))
            c6.border = self.styles.border; c6.alignment = self.styles.center_align
            
            fila += 1
            
        if fila > fila_h + 1:
            rule = DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color="5A8AD6", showValue="None", minLength=None, maxLength=None)
            ws.conditional_formatting.add(f"E{fila_h+1}:E{fila-1}", rule)

        # Autoajuste de columnas
        self._auto_adjust_columns(ws)

        # Añadir gráfico de columnas para Top Clientes
        try:
            from openpyxl.chart import BarChart, Reference
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
            chart.title = "Top 10 Clientes"
            chart.y_axis.title = 'Monto S/'
            chart.x_axis.title = 'Vendedor / Cliente'
            
            # Tomar los top 10 clientes de la tabla que acabamos de escribir
            num_clients = min(len(clientes), 10)
            if num_clients > 0:
                # data: incluye cabecera "TOTAL SOLES" como título de la serie + num_clients filas de datos
                data = Reference(ws, min_col=4, min_row=fila_h, max_row=fila_h+num_clients)
                # categories: CLIENTE + VENDEDOR (col 2-3), una fila por barra
                cats = Reference(ws, min_col=2, min_row=fila_h+1, max_col=3, max_row=fila_h+num_clients)
                
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(cats)
                chart.shape = 4
                chart.height = 15.6  # Cubre hasta la fila 30 (aprox 15.6 cm)
                chart.width = 24     # Cubre hasta la columna T (aprox 24 cm)
                
                # Ajustar tamaño y posicionar en I1 (I1:U25)
                chart.width = 28
                chart.height = 18
                ws.add_chart(chart, "I1")
        except Exception:
            pass

        # === LEYENDA / CARTILLA DE INTERPRETACIÓN ===
        fila_leyenda = fila + 2
        ws.cell(row=fila_leyenda, column=1, value="CÓMO LEER ESTE REPORTE").font = Font(bold=True, size=11, color="0D2B4E")
        fila_leyenda += 1
        leyendas = [
            ("ICONOS", "🔺 Crecimiento (>5%)  |  🔻 Caída (>5%)  |  ➡️ Estable  |  ⏳ Mes en proceso"),
            ("COLORES", "🔴 Rojo = Alerta / Atención  |  🟡 Ámbar = Precaución  |  🟢 Verde = Saludable"),
            ("CATEGORÍAS", "VITAL (≤80%) = Clientes que generan el 80% de la facturación  |  TRIVIAL (>80%) = Clientes menores"),
            ("VULNERABILIDAD", "ALTA = >80% de su compra en una sola línea  |  MEDIA = 60-80%  |  BAJA = <60% (diversificado)"),
            ("HHI (KPI)", "<0.15 = Baja concentración  |  0.15-0.25 = Moderada  |  >0.25 = ALTA (riesgo de dependencia)"),
            ("DEV%", "Devoluciones / Venta Total. >5% = Alerta (impacta comisión neta)"),
            ("FÓRMULAS", "Las celdas de % contienen fórmulas de Excel. Use la fila TOTAL (SUBTOTAL) para filtrar por categoría."),
        ]
        for label, text in leyendas:
            ws.cell(row=fila_leyenda, column=1, value=label).font = Font(bold=True, size=9, color="0D2B4E")
            ws.cell(row=fila_leyenda, column=1).border = self.styles.border
            ws.cell(row=fila_leyenda, column=2, value=text).font = Font(size=9)
            ws.cell(row=fila_leyenda, column=2).border = self.styles.border
            ws.merge_cells(start_row=fila_leyenda, start_column=2, end_row=fila_leyenda, end_column=8)
            fila_leyenda += 1

        self._auto_adjust_columns(ws)

    def _escribir_anexo_sucursales(self, datos_sucursales, historial, orden_vendedores=None):
        """Anexo Pareto por Cliente + Sucursal, una hoja por vendedor.
        Solo muestra clientes con más de 1 sucursal."""
        from openpyxl.utils import get_column_letter
        # Determinar orden: usar el mismo de POR_VENDEDOR si está disponible
        if orden_vendedores:
            # Obtener solo los vendedores que tienen datos de sucursales, en el orden de POR_VENDEDOR
            claves_orden = []
            for k in orden_vendedores.keys():
                # Extraer ID del vendedor del display "ID - NOMBRE"
                vid = k.split(' - ')[0] if ' - ' in k else (k.split('-')[0] if '-' in k else k)
                if vid in datos_sucursales:
                    claves_orden.append(vid)
            # Agregar vendedores que estén en sucursales pero no en el orden
            for k in datos_sucursales:
                if k not in claves_orden:
                    claves_orden.append(k)
        else:
            claves_orden = list(datos_sucursales.keys())
        
        # Para cada vendedor, filtrar solo clientes con +1 sucursal
        for vendedor_id in claves_orden:
            items = datos_sucursales.get(vendedor_id, [])
            if not items:
                continue
            
            # Contar sucursales por cliente y filtrar solo multicanal (+1 sucursal)
            from collections import Counter
            # Extraer cliente (ID - NOMBRE) de cada item
            clientes_sucursales = Counter()
            for it in items:
                cli_key = it.get('CLIENTE', '')
                clientes_sucursales[cli_key] += 1
            
            clientes_multisucursal = {c for c, cnt in clientes_sucursales.items() if cnt > 1}
            
            # Si ningún cliente tiene +1 sucursal, saltar este vendedor
            if not clientes_multisucursal:
                continue
            
            items_filtrados = [it for it in items if it.get('CLIENTE', '') in clientes_multisucursal]
            
            # Reordenar para congruencia: clientes por monto total descendente (mismo criterio Pareto Líneas)
            # Agrupar items por cliente y sumar montos
            montos_por_cliente = {}
            for it in items_filtrados:
                cli = it.get('CLIENTE', '')
                montos_por_cliente[cli] = montos_por_cliente.get(cli, 0) + it.get('MONTO', 0)
            # Orden clientes por monto total descendente
            orden_clientes = sorted(montos_por_cliente.keys(), key=lambda c: montos_por_cliente[c], reverse=True)
            # Reordenar items_filtrados según ese orden y dentro de cada cliente por monto descendente
            items_ordenados = []
            for cli in orden_clientes:
                items_cli = [it for it in items_filtrados if it.get('CLIENTE', '') == cli]
                items_cli.sort(key=lambda x: x.get('MONTO', 0), reverse=True)
                items_ordenados.extend(items_cli)
            items_filtrados = items_ordenados
            
            # Obtener nombre del vendedor del historial
            nom_vendedor = ""
            if historial is not None and 'NOM_VENDEDOR' in historial.columns:
                df_ven = historial[historial['ID_VENDEDOR'] == vendedor_id]
                if not df_ven.empty:
                    nom_vendedor = df_ven['NOM_VENDEDOR'].iloc[0]
            primer_palabra = nom_vendedor.split(' ')[0] if nom_vendedor else vendedor_id
            sheet_name = f"SUC_{vendedor_id}_{primer_palabra}"[:31]
            sheet_name = re.sub(r'[\\/*?:\[\]]', "", sheet_name)
            if sheet_name in self.wb.sheetnames:
                sheet_name = f"{sheet_name}_c"[:31]
            
            ws = self.wb.create_sheet(sheet_name)
            ws.cell(row=1, column=1, value=f"VENTAS POR CLIENTE + SUCURSAL: {format_id_name(vendedor_id, nom_vendedor, field_name='VENDEDOR')}").font = self.styles.title_font
            
            cabeceras = ['N°', 'CLIENTE', 'SUCURSAL', 'MONTO TOTAL', 'CANT', 'FACTURAS', '% IND', '% ACUM', 'CAT']
            for i, text in enumerate(cabeceras, 1):
                c = ws.cell(row=4, column=i, value=text)
                c.fill = self.styles.header_fill; c.font = self.styles.header_font; c.border = self.styles.border; c.alignment = self.styles.center_align
            
            fila = 5
            for idx, item in enumerate(items_filtrados):
                nro = idx + 1
                ws.cell(row=fila, column=1, value=nro).border = self.styles.border; ws.cell(row=fila, column=1).alignment = self.styles.center_align
                ws.cell(row=fila, column=2, value=item.get('CLIENTE', '')).border = self.styles.border
                ws.cell(row=fila, column=3, value=item.get('SUCURSAL', '')).border = self.styles.border
                ws.cell(row=fila, column=4, value=item.get('MONTO', 0)).number_format = '#,##0.00'; ws.cell(row=fila, column=4).border = self.styles.border
                ws.cell(row=fila, column=5, value=item.get('CANTIDAD', 0)).border = self.styles.border; ws.cell(row=fila, column=5).alignment = self.styles.center_align
                ws.cell(row=fila, column=6, value=item.get('FACTURAS', '')).border = self.styles.border
                # CORRECCIÓN: PCT_INDIVIDUAL viene en porcentaje, dividir entre 100 para formato decimal de Excel
                pct_ind = item.get('PCT_INDIVIDUAL', 0) / 100
                ws.cell(row=fila, column=7, value=pct_ind).number_format = '0.00%'; ws.cell(row=fila, column=7).border = self.styles.border
                # CORRECCIÓN: PCT_ACUMULADO viene en porcentaje, dividir entre 100 para formato decimal de Excel
                pct_acum = item.get('PCT_ACUMULADO', 0) / 100
                ws.cell(row=fila, column=8, value=pct_acum).number_format = '0.00%'; ws.cell(row=fila, column=8).border = self.styles.border
                ws.cell(row=fila, column=9, value=item.get('CATEGORIA', '')).border = self.styles.border; ws.cell(row=fila, column=9).alignment = self.styles.center_align
                fila += 1
            
            # DataBar para % Individual
            if items:
                rule = DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color="5A8AD6", showValue="None")
                ws.conditional_formatting.add(f"G5:G{fila-1}", rule)
            
            self._auto_adjust_columns(ws)
            ws.freeze_panes = "C5"

    def _escribir_anexo_evolucion(self, datos_pareto, historial, rango_fechas, orden_vendedores=None):
        """Hojas para medir los meses presentes en el historial (Análisis de Evolución por Vendedor).
        Usa orden_vendedores para mantener congruencia con el orden de Pareto Líneas."""
        periodos = datos_pareto.get('PERIODOS', [])
        
        # Agrupar historial por Cliente y Mes
        if 'PERIODO_TEND' not in historial.columns:
            historial['PERIODO_TEND'] = historial['FECHA_ORIG'].dt.to_period('M').astype(str)
            
        if 'ID_VENDEDOR' in historial.columns and 'NOM_VENDEDOR' in historial.columns:
            df_mensual = historial.groupby(['ID_VENDEDOR', 'NOM_VENDEDOR', 'ID_CLIENTE', 'NOM_CLIENTE', 'PERIODO_TEND'])['SOLES'].sum().unstack(fill_value=0)
            usar_id_vendedor = True
        else:
            df_mensual = historial.groupby(['NOM_VENDEDOR', 'ID_CLIENTE', 'NOM_CLIENTE', 'PERIODO_TEND'])['SOLES'].sum().unstack(fill_value=0)
            usar_id_vendedor = False
            
        # Obtener los vendedores únicos
        if usar_id_vendedor:
            vendedores_keys = df_mensual.index.droplevel(['ID_CLIENTE', 'NOM_CLIENTE']).unique()
        else:
            vendedores_keys = df_mensual.index.droplevel(['ID_CLIENTE', 'NOM_CLIENTE']).unique()
        
        # Reordenar para mantener congruencia con Pareto Líneas si hay orden_vendedores
        if orden_vendedores and usar_id_vendedor:
            # Extraer IDs de vendedor del orden_vendedores (formato "ID - NOMBRE")
            orden_ids = {}
            for v_display in orden_vendedores.keys():
                partes = v_display.split(' - ') if ' - ' in v_display else (v_display.split('-') if '-' in v_display else [v_display])
                vid = partes[0] if partes else v_display
                if ' - ' in v_display:
                    vnom = v_display.split(' - ', 1)[1]
                elif '-' in v_display:
                    vnom = v_display.split('-', 1)[1]
                else:
                    vnom = ""
                orden_ids[vid] = vnom
            # Reordenar vendedores_keys según el orden de orden_vendedores
            vendedores_ordenados = []
            for vid_orden, _ in orden_ids.items():
                for vk in vendedores_keys:
                    if vk[0] == vid_orden:
                        vendedores_ordenados.append(vk)
                        break
            # Agregar los que no estaban en orden
            for vk in vendedores_keys:
                if vk not in vendedores_ordenados:
                    vendedores_ordenados.append(vk)
            vendedores_keys = vendedores_ordenados
            
        # Eliminar la hoja "EVOLUCION_MENSUAL" si existiera previamente
        if "EVOLUCION_MENSUAL" in self.wb.sheetnames:
            del self.wb["EVOLUCION_MENSUAL"]
            
        for v_key in vendedores_keys:
            if usar_id_vendedor:
                id_v, nom_v = v_key
                vendedor_display = format_id_name(id_v, nom_v)
                df_vendedor = df_mensual.loc[(id_v, nom_v)]
            else:
                nom_v = v_key
                vendedor_display = str(nom_v)
                df_vendedor = df_mensual.loc[nom_v]
                
            # Crear nombre de hoja seguro
            sheet_name = f"EVOL_{vendedor_display[:26]}"[:31]
            import re
            sheet_name = re.sub(r'[\\/*?:\[\]]', "", sheet_name)
            
            # Evitar duplicados
            if sheet_name in self.wb.sheetnames:
                sheet_name = f"{sheet_name}_c"[:31]
                
            ws = self.wb.create_sheet(sheet_name)
            ws.cell(row=1, column=1, value=f"EVOLUCIÓN MENSUAL: {vendedor_display}").font = self.styles.title_font
            
            cabeceras = ['N°', 'CLIENTE'] + [p.upper() for p in periodos] + ['PROM. MENSUAL']
            for i, text in enumerate(cabeceras, 1):
                c = ws.cell(row=4, column=i, value=text)
                c.fill = self.styles.header_fill; c.font = self.styles.header_font; c.border = self.styles.border
                
            # Calcular total por cliente para ordenar (mismo criterio que Pareto Líneas: monto descendente)
            clientes_totales = []
            for row_key, row_vals in df_vendedor.iterrows():
                total = sum(row_vals.get(p, 0) for p in periodos)
                clientes_totales.append((row_key, row_vals, total))
            # Ordenar por total descendente (mayor monto primero)
            clientes_totales.sort(key=lambda x: x[2], reverse=True)
            
            fila = 5
            for idx, (row_key, row_vals, _) in enumerate(clientes_totales):
                id_c, nom_c = row_key
                nro = idx + 1
                ws.cell(row=fila, column=1, value=nro).border = self.styles.border; ws.cell(row=fila, column=1).alignment = self.styles.center_align
                ws.cell(row=fila, column=2, value=format_id_name(id_c, nom_c)).border = self.styles.border
                
                c_vals = 3
                sum_soles = 0
                previous_month_value = 0
                
                for p in periodos:
                    val = row_vals.get(p, 0)
                    sum_soles += val
                    cell = ws.cell(row=fila, column=c_vals, value=val)
                    cell.number_format = '#,##0.00'; cell.border = self.styles.border
                    
                    if previous_month_value > 0 and (val - previous_month_value) / previous_month_value < -0.20:
                        cell.fill = self.styles.warning_fill
                        cell.font = self.styles.alert_font
                        
                    previous_month_value = val
                    c_vals += 1
                
                prom = ws.cell(row=fila, column=c_vals, value=sum_soles / len(periodos) if periodos else 0)
                prom.number_format = '#,##0.00'; prom.border = self.styles.border; prom.fill = self.styles.total_fill
                fila += 1
                
        self._auto_adjust_columns(ws)
        ws.freeze_panes = "B5"

    def _escribir_hoja_vendedor_integrada(self, ws, vendedor_display, clientes, lineas, historial, datos_sucursales, rango_fechas):
        """Genera hoja integrada por vendedor con flujo vertical: Header KPI + 4 tablas."""
        from openpyxl.utils import get_column_letter
        from openpyxl.formatting.rule import DataBarRule
        from openpyxl.chart import BarChart, Reference
        
        # Extraer ID y nombre del vendedor
        partes = vendedor_display.split(' - ') if ' - ' in vendedor_display else (vendedor_display.split('-') if '-' in vendedor_display else [vendedor_display])
        id_vendedor = partes[0] if partes else vendedor_display
        nom_vendedor = partes[1] if len(partes) > 1 else ""
        
        # Fila 1: Título del vendedor
        ws.cell(row=1, column=1, value=f"VENDEDOR: {vendedor_display}").font = self.styles.title_font
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
        
        # ============================================================
        # RECALCULAR % PARA EL UNIVERSO DEL VENDEDOR
        # ============================================================
        # Los datos vienen con PCT_GLOBAL calculado contra el total de la empresa.
        # Para la hoja del vendedor, debemos recalcular contra SU propio total.
        total_facturacion = sum(c.get('MONTO_TOTAL', 0) for c in clientes)
        if total_facturacion > 0:
            for cliente in clientes:
                monto = cliente.get('MONTO_TOTAL', 0)
                pct_recalc = (monto / total_facturacion) * 100
                cliente['PCT_GLOBAL'] = round(pct_recalc, 2)
        
        # Reordenar por monto descendente (por si acaso)
        clientes.sort(key=lambda c: c.get('MONTO_TOTAL', 0), reverse=True)
        
        # Recalcular % ACUM y CATEGORÍA para el universo del vendedor
        acum = 0
        for cliente in clientes:
            acum += cliente.get('PCT_GLOBAL', 0)
            cliente['PCT_ACUM'] = round(acum, 2)
            if acum <= 80:
                cliente['CATEGORIA'] = 'VITAL (≤80%)'
            elif acum == 100:
                cliente['CATEGORIA'] = 'VITAL (100%)'
            else:
                cliente['CATEGORIA'] = 'TRIVIAL (>80%)'
        
        # Calcular KPIs del universo del vendedor
        total_clientes = len(clientes)
        clientes_vitales = sum(1 for c in clientes if 'VITAL' in str(c.get('CATEGORIA', '')).upper())
        
        # Índice de Dependencia: % del cliente #1 sobre el total
        # CORRECCIÓN: El valor ya está en porcentaje (ej. 14.96), pero Excel con formato "0.00%" multiplica por 100
        # Solución: Dividir entre 100 para formato decimal de Excel
        indice_dependencia = 0
        if clientes and total_facturacion > 0:
            cliente_1_monto = clientes[0].get('MONTO_TOTAL', 0)
            indice_dependencia = (cliente_1_monto / total_facturacion)  # Valor decimal (ej. 0.1496)
        
        # FASE 2: Índice de Dispersión (Clientes Triviales / Venta Trivial)
        clientes_triviales = [c for c in clientes if 'VITAL' not in str(c.get('CATEGORIA', '')).upper()]
        venta_trivial = sum(c.get('MONTO_TOTAL', 0) for c in clientes_triviales)
        num_triviales = len(clientes_triviales)
        indice_dispersion = (num_triviales / venta_trivial) if venta_trivial > 0 else 0
        
        # SUGERENCIA 1: HHI (Índice de Concentración Herfindahl-Hirschman)
        # HHI = Σ(market_share_i²)
        hhi = 0
        if total_facturacion > 0:
            hhi = sum((c.get('MONTO_TOTAL', 0) / total_facturacion) ** 2 for c in clientes)
        
        # === NOTAS DE CRÉDITO (devoluciones) ===
        total_devoluciones = 0
        devoluciones_por_cliente = {}
        devoluciones_por_periodo = {}
        if historial is not None and 'SOLES' in historial.columns:
            df_dev = historial[historial['SOLES'] < 0].copy()
            if not df_dev.empty and 'ID_VENDEDOR' in df_dev.columns:
                df_dev = df_dev[df_dev['ID_VENDEDOR'].astype(str) == str(id_vendedor)]
                if not df_dev.empty:
                    if rango_fechas and rango_fechas[0] and rango_fechas[1]:
                        df_dev = df_dev[(df_dev['FECHA_ORIG'] >= rango_fechas[0]) & (df_dev['FECHA_ORIG'] <= rango_fechas[1])]
                    if not df_dev.empty:
                        if 'PERIODO_TEND' not in df_dev.columns:
                            df_dev['PERIODO_TEND'] = df_dev['FECHA_ORIG'].dt.to_period('M').astype(str)
                        total_devoluciones = abs(df_dev['SOLES'].sum())
                        dev_cliente = df_dev.groupby('ID_CLIENTE')['SOLES'].sum().abs()
                        devoluciones_por_cliente = {str(k).strip(): v for k, v in dev_cliente.items()}
                        dev_periodo = df_dev.groupby('PERIODO_TEND')['SOLES'].sum().abs()
                        devoluciones_por_periodo = dev_periodo.to_dict()
        
        # Fila 2: Header KPI
        dev_pct = (total_devoluciones / total_facturacion) if total_facturacion > 0 else 0
        kpi_labels = ["TOTAL S/", "CLIENTES", "VITALES", "DEPENDENCIA", "DISPERSIÓN", "HHI", "DEVOLUCIONES", "DEV%"]
        kpi_values = [total_facturacion, total_clientes, clientes_vitales, indice_dependencia, indice_dispersion, hhi, total_devoluciones, dev_pct]
        kpi_formats = ["#,##0.00", "#,##0", "#,##0", "0.00%", "0.0000", "0.0000", "#,##0.00", "0.00%"]
        
        for i, (label, val, fmt) in enumerate(zip(kpi_labels, kpi_values, kpi_formats), 1):
            c_label = ws.cell(row=2, column=i, value=label)
            c_label.font = self.styles.kpi_label_font
            c_label.fill = self.styles.kpi_header_fill
            c_label.border = self.styles.border
            c_label.alignment = self.styles.center_align
            
            c_value = ws.cell(row=3, column=i, value=val)
            c_value.font = self.styles.kpi_value_font
            c_value.number_format = fmt
            c_value.fill = self.styles.kpi_value_fill
            c_value.border = self.styles.border
            c_value.alignment = self.styles.center_align
        
        # Alarma HHI
        c_hhi = ws.cell(row=3, column=6)
        if hhi > 0.25:
            c_hhi.fill = self.styles.hhi_high_fill
            c_hhi.font = self.styles.hhi_high_font
        elif hhi >= 0.15:
            c_hhi.fill = self.styles.hhi_mod_fill
            c_hhi.font = self.styles.hhi_mod_font
        else:
            c_hhi.fill = self.styles.hhi_low_fill
            c_hhi.font = self.styles.hhi_low_font
        
        # Alarma DEV%: >5% de devoluciones sobre venta es alto
        c_devpct = ws.cell(row=3, column=8)
        if dev_pct > 0.05:
            c_devpct.fill = self.styles.hhi_mod_fill
            c_devpct.font = self.styles.hhi_mod_font
        
        # Fila 4: Espacio
        ws.cell(row=4, column=1, value="")
        
        # === TABLA A: PARETO GLOBAL ===
        fila_inicio_pareto = 5
        ws.cell(row=fila_inicio_pareto, column=1, value="TABLA A: PARETO GLOBAL").font = Font(bold=True, size=11)
        
        fila_cab_pareto = fila_inicio_pareto + 1
        cabeceras_pareto = ['N°', 'CLIENTE', 'TOTAL SOLES', '% GLOB', '% ACUM', 'TEND', 'CAT', 'VULNERAB', 'DEVOLUCIONES S/']
        for i, text in enumerate(cabeceras_pareto, 1):
            c = ws.cell(row=fila_cab_pareto, column=i, value=text)
            c.fill = self.styles.header_fill
            c.font = self.styles.header_font
            c.border = self.styles.border
            c.alignment = self.styles.center_align
        
        # Calcular vulnerabilidad para cada cliente
        for cliente in clientes:
            cliente_total = cliente.get('MONTO_TOTAL', 0)
            if cliente_total == 0:
                cliente['VULNERABILIDAD'] = 'N/A'
                continue
            
            # Encontrar la línea con mayor participación
            max_linea_pct = 0
            for linea in lineas:
                lid = linea['ID_LINEA']
                linea_monto = cliente.get(f'L{lid}_MONTO', 0)
                linea_pct = (linea_monto / cliente_total) * 100 if cliente_total > 0 else 0
                if linea_pct > max_linea_pct:
                    max_linea_pct = linea_pct
            
            # Clasificar vulnerabilidad
            if max_linea_pct > 80:
                cliente['VULNERABILIDAD'] = 'ALTA'
            elif max_linea_pct >= 60:
                cliente['VULNERABILIDAD'] = 'MEDIA'
            else:
                cliente['VULNERABILIDAD'] = 'BAJA'
        
        # Escribir datos de Pareto
        fila_datos_pareto = fila_cab_pareto + 1
        for idx, cli in enumerate(clientes):
            es_vital = 'VITAL' in str(cli.get('CATEGORIA', '')).upper()
            nro = idx + 1
            
            # Col A: N°
            c0 = ws.cell(row=fila_datos_pareto, column=1, value=nro)
            c0.border = self.styles.border
            c0.alignment = self.styles.center_align
            
            # Col B: CLIENTE (left)
            c1 = ws.cell(row=fila_datos_pareto, column=2, value=cli.get('CLIENTE', ''))
            c1.border = self.styles.border
            c1.alignment = self.styles.left_align
            
            # Col C: TOTAL SOLES (right)
            c2 = ws.cell(row=fila_datos_pareto, column=3, value=cli.get('MONTO_TOTAL', 0))
            c2.number_format = '#,##0.00'
            c2.border = self.styles.border
            c2.alignment = self.styles.right_align
            
            # Col D: % GLOB
            # CORRECCIÓN: PCT_GLOBAL viene en porcentaje (ej. 4.07, 14.96), dividir entre 100 para formato decimal de Excel
            pct_global = cli.get('PCT_GLOBAL', 0) / 100
            c3 = ws.cell(row=fila_datos_pareto, column=4, value=pct_global)
            c3.number_format = '0.00%'
            c3.border = self.styles.border
            
            # Col E: % ACUM (recalculado para el universo del vendedor)
            pct_acum = cli.get('PCT_ACUM', 0) / 100
            c4 = ws.cell(row=fila_datos_pareto, column=5, value=pct_acum)
            c4.number_format = '0.00%'
            c4.border = self.styles.border
            
            # Col F: TEND
            c5 = ws.cell(row=fila_datos_pareto, column=6, value=cli.get('TENDENCIA', '➡️'))
            c5.alignment = self.styles.center_align
            c5.border = self.styles.border
            if '🔺' in str(c5.value):
                c5.font = self.styles.trend_up_font
            elif '🔻' in str(c5.value):
                c5.font = self.styles.trend_down_font
                # Alerta de tendencia en clientes VITALES
                if es_vital:
                    c5.fill = self.styles.alert_fill
            
            # Col G: CAT
            c6 = ws.cell(row=fila_datos_pareto, column=7, value=cli.get('CATEGORIA', ''))
            c6.border = self.styles.border
            c6.alignment = self.styles.center_align
            
            # Col H: VULNERAB
            c7 = ws.cell(row=fila_datos_pareto, column=8, value=cli.get('VULNERABILIDAD', 'N/A'))
            c7.border = self.styles.border
            c7.alignment = self.styles.center_align
            if c7.value == 'ALTA':
                c7.font = self.styles.vuln_high_font
            elif c7.value == 'MEDIA':
                c7.font = self.styles.vuln_mod_font
            
            # Col I: DEVOLUCIONES S/ (right)
            cli_id_dev = str(cli.get('ID_CLIENTE', '')).strip()
            dev_monto = devoluciones_por_cliente.get(cli_id_dev, 0)
            c8 = ws.cell(row=fila_datos_pareto, column=9, value=dev_monto)
            c8.number_format = '#,##0.00'
            c8.border = self.styles.border
            c8.alignment = self.styles.right_align
            if dev_monto > 0:
                c8.font = self.styles.dev_font
            
            fila_datos_pareto += 1
        
        # === CREAR TABLA EXCEL + FÓRMULAS ESTRUCTURADAS (Tabla A) ===
        if clientes:
            first_data = fila_cab_pareto + 1
            last_data = fila_datos_pareto - 1
            subtotal_row = fila_datos_pareto
            
            from openpyxl.worksheet.table import Table, TableStyleInfo
            import re
            safe_name = re.sub(r'[^A-Za-z0-9]', '', ws.title)[:10]
            table_name = f"TblA_{safe_name}"
            
            # Crear tabla Excel sobre el rango A{header}:I{last_data}
            tabla_ref = f"A{fila_cab_pareto}:I{last_data}"
            tab = Table(displayName=table_name, ref=tabla_ref)
            tab.tableStyleInfo = TableStyleInfo(name="TableStyleLight9", showRowStripes=True)
            ws.add_table(tab)
            
            # SUBTOTAL(9,) en col C referenciando la tabla
            c_sub = ws.cell(row=subtotal_row, column=3)
            c_sub.value = f'=SUBTOTAL(9,{table_name}[TOTAL SOLES])'
            c_sub.number_format = '#,##0.00'
            c_sub.border = self.styles.border
            c_sub.font = self.styles.subtotal_font
            c_sub.fill = self.styles.subtotal_fill
            ws.cell(row=subtotal_row, column=1, value='').border = self.styles.border
            ws.cell(row=subtotal_row, column=2, value='TOTAL').border = self.styles.border
            ws.cell(row=subtotal_row, column=2).font = self.styles.subtotal_font
            ws.cell(row=subtotal_row, column=2).alignment = self.styles.center_align
            for sc in range(4, 10):
                ws.cell(row=subtotal_row, column=sc).border = self.styles.border
            # SUBTOTAL DEVOLUCIONES
            c_dev_sub = ws.cell(row=subtotal_row, column=9)
            c_dev_sub.value = f'=SUBTOTAL(9,{table_name}[DEVOLUCIONES S/])'
            c_dev_sub.number_format = '#,##0.00'
            c_dev_sub.font = self.styles.dev_sub_font
            c_dev_sub.fill = self.styles.subtotal_fill
            
            # % GLOB y % ACUM con celdas simples (compatible con todas las versiones de Excel)
            for r in range(first_data, last_data + 1):
                ws.cell(row=r, column=4).value = f'=IF(C{r}=0,0,C{r}/C${subtotal_row})'
                if r == first_data:
                    ws.cell(row=r, column=5).value = f'=D{r}'
                else:
                    ws.cell(row=r, column=5).value = f'=E{r-1}+D{r}'
        
        # DataBar para % GLOB (sobre fórmulas, Excel evalúa el resultado)
        if clientes:
            rule = DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color="5A8AD6", showValue="None")
            ws.conditional_formatting.add(f"D{first_data}:D{last_data}", rule)
        
        # === GRÁFICO TOP 10 (simple, no apilado) ===
        try:
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
            chart.title = "Top 10 Clientes"
            chart.y_axis.title = 'Monto S/'
            chart.x_axis.title = 'Cliente'
            
            num_clients = min(len(clientes), 10)
            if num_clients > 0:
                # Data: TOTAL SOLES (col C) — incluye cabecera como título + num_clients barras
                data = Reference(ws, min_col=3, min_row=fila_cab_pareto, max_row=fila_cab_pareto+num_clients)
                # Categories: CLIENTE (col B), una por barra
                cats = Reference(ws, min_col=2, min_row=fila_cab_pareto+1, max_row=fila_cab_pareto+num_clients)
                
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(cats)
                chart.shape = 4
                chart.width = 20
                chart.height = 12
                
                # Posicionar gráfico a la derecha de la Tabla A (2 columnas de margen)
                chart_col = get_column_letter(len(cabeceras_pareto) + 2)
                ws.add_chart(chart, f"{chart_col}5")
        except Exception:
            pass
        
        # === TABLA B: DESGLOSE POR LÍNEAS (columnas duales: Monto + %) ===
        fila_inicio_lineas = fila_datos_pareto + 2
        ws.cell(row=fila_inicio_lineas, column=1, value="TABLA B: DESGLOSE POR LÍNEAS (Monto + %)").font = Font(bold=True, size=11)
        
        fila_cab_lineas = fila_inicio_lineas + 1
        # Cabeceras base
        cabeceras_lineas = ['N°', 'CLIENTE']
        for linea in lineas:
            lid = linea['ID_LINEA']
            lnom = linea.get('NOM_LINEA', '')
            linea_display = format_id_name(lid, lnom, field_name='LÍNEA')
            cabeceras_lineas.append(f"{linea_display} S/")
            cabeceras_lineas.append(f"{linea_display} %")
        cabeceras_lineas.append('TOTAL SOLES')
        
        for i, text in enumerate(cabeceras_lineas, 1):
            c = ws.cell(row=fila_cab_lineas, column=i, value=text)
            c.fill = self.styles.header_fill
            c.font = self.styles.header_font
            c.border = self.styles.border
            c.alignment = self.styles.center_align
        
        # FASE 3: Calcular líneas que compran los clientes VITALES (para alerta de Línea Blanca)
        lineas_compradas_por_vitales = set()
        monto_linea_en_vitales = {}
        for cli in clientes:
            if 'VITAL' in str(cli.get('CATEGORIA', '')).upper():
                for linea in lineas:
                    lid = linea['ID_LINEA']
                    linea_monto = cli.get(f'L{lid}_MONTO', 0)
                    if linea_monto > 0:
                        lineas_compradas_por_vitales.add(lid)
                        monto_linea_en_vitales[lid] = monto_linea_en_vitales.get(lid, 0) + linea_monto
        
        # Columna TOTAL SOLES al final de Tabla B
        has_total_col = True
        
        # Escribir datos de líneas
        fila_datos_lineas = fila_cab_lineas + 1
        for idx, cli in enumerate(clientes):
            cliente_total = cli.get('MONTO_TOTAL', 0)
            nro = idx + 1
            es_vital = 'VITAL' in str(cli.get('CATEGORIA', '')).upper()
            
            # Col A: N°
            c0 = ws.cell(row=fila_datos_lineas, column=1, value=nro)
            c0.border = self.styles.border
            c0.alignment = self.styles.center_align
            
            # Col B: CLIENTE (left)
            c1 = ws.cell(row=fila_datos_lineas, column=2, value=cli.get('CLIENTE', ''))
            c1.border = self.styles.border
            c1.alignment = self.styles.left_align
            
            # Columnas de líneas (Monto + %) - right align
            col_idx = 3
            for linea in lineas:
                lid = linea['ID_LINEA']
                linea_monto = cli.get(f'L{lid}_MONTO', 0)
                linea_pct = (linea_monto / cliente_total) * 100 if cliente_total > 0 else 0
                
                # Columna de Monto (right)
                c_monto = ws.cell(row=fila_datos_lineas, column=col_idx, value=linea_monto)
                c_monto.number_format = '#,##0.00'
                c_monto.border = self.styles.border
                c_monto.alignment = self.styles.right_align
                col_idx += 1
                
                # Columna de % (right)
                c_pct = ws.cell(row=fila_datos_lineas, column=col_idx, value=linea_pct / 100)
                c_pct.number_format = '0.00%'
                c_pct.border = self.styles.border
                c_pct.alignment = self.styles.right_align
                
                # FASE 3: Alerta de Línea Blanca (venta cruzada)
                # Si es cliente VITAL, tiene 0% en esta línea, otros VITALES compran, y la línea mueve > S/50
                if es_vital and linea_pct == 0 and lid in lineas_compradas_por_vitales and monto_linea_en_vitales.get(lid, 0) >= 50:
                    c_pct.fill = self.styles.alert_fill
                    c_pct.font = self.styles.alert_font
                
                col_idx += 1
            
            # Columna TOTAL SOLES (fórmula SUM sobre columnas de monto, right)
            amount_cols = [c for c in range(3, col_idx, 2)]
            formula_parts = [f'{get_column_letter(c)}{fila_datos_lineas}' for c in amount_cols]
            formula_total = '=' + '+'.join(formula_parts)
            c_total = ws.cell(row=fila_datos_lineas, column=col_idx, value=formula_total)
            c_total.number_format = '#,##0.00'
            c_total.border = self.styles.border
            c_total.alignment = self.styles.right_align
            c_total.font = self.styles.bold_font
            
            fila_datos_lineas += 1
        
        # SUBTOTAL fila en Tabla B (solo columna TOTAL SOLES)
        if clientes:
            b_first = fila_cab_lineas + 1
            b_last = fila_datos_lineas - 1
            b_sub = fila_datos_lineas
            total_col_letter = get_column_letter(3 + 2 * len(lineas))  # columna TOTAL SOLES
            for bc in range(1, 3 + 2 * len(lineas) + 1):
                ws.cell(row=b_sub, column=bc).border = self.styles.border
            ws.cell(row=b_sub, column=1).value = ''
            ws.cell(row=b_sub, column=2, value='TOTAL').font = self.styles.subtotal_font
            ws.cell(row=b_sub, column=2).alignment = self.styles.center_align
            c_bsub = ws.cell(row=b_sub, column=3 + 2 * len(lineas))
            c_bsub.value = f'=SUBTOTAL(9, {total_col_letter}{b_first}:{total_col_letter}{b_last})'
            c_bsub.number_format = '#,##0.00'
            c_bsub.font = self.styles.subtotal_font
            c_bsub.fill = self.styles.subtotal_fill
            fila_datos_lineas = b_sub + 1  # avanzar para próxima tabla
        
        # === TABLA C: SUCURSALES (solo multicanal) ===
        if datos_sucursales:
            # Filtrar solo clientes con más de 1 sucursal
            from collections import Counter
            clientes_sucursales = Counter()
            for it in datos_sucursales:
                cli_key = it.get('CLIENTE', '')
                clientes_sucursales[cli_key] += 1
            
            clientes_multisucursal = {c for c, cnt in clientes_sucursales.items() if cnt > 1}
            
            if clientes_multisucursal:
                items_filtrados = [it for it in datos_sucursales if it.get('CLIENTE', '') in clientes_multisucursal]
                
                # FIX 5: Ordenar sucursales por el orden del Pareto del vendedor
                orden_pareto = {cli.get('CLIENTE', ''): idx for idx, cli in enumerate(clientes)}
                items_filtrados.sort(key=lambda x: orden_pareto.get(x.get('CLIENTE', ''), 9999))
                
                if items_filtrados:
                    fila_inicio_sucursales = fila_datos_lineas + 2
                    ws.cell(row=fila_inicio_sucursales, column=1, value="TABLA C: SUCURSALES (multicanal)").font = Font(bold=True, size=11)
                    
                    fila_cab_sucursales = fila_inicio_sucursales + 1
                    # SUGERENCIA 3: Agregar DROP SIZE y FRECUENCIA
                    cabeceras_sucursales = ['N°', 'CLIENTE', 'SUCURSAL', 'MONTO', 'CANT', 'FACTURAS', 'DROP SIZE', 'FRECUENCIA', '%', 'CAT']
                    for i, text in enumerate(cabeceras_sucursales, 1):
                        c = ws.cell(row=fila_cab_sucursales, column=i, value=text)
                        c.fill = self.styles.header_fill
                        c.font = self.styles.header_font
                        c.border = self.styles.border
                        c.alignment = self.styles.center_align
                    
                    # Calcular número de meses para frecuencia
                    num_meses = len(periodos) if 'periodos' in locals() else 1
                    
                    # Escribir datos de sucursales
                    fila_datos_sucursales = fila_cab_sucursales + 1
                    for idx, item in enumerate(items_filtrados):
                        nro = idx + 1
                        ws.cell(row=fila_datos_sucursales, column=1, value=nro).border = self.styles.border
                        ws.cell(row=fila_datos_sucursales, column=1).alignment = self.styles.center_align
                        ws.cell(row=fila_datos_sucursales, column=2, value=item.get('CLIENTE', '')).border = self.styles.border
                        ws.cell(row=fila_datos_sucursales, column=3, value=item.get('SUCURSAL', '')).border = self.styles.border
                        ws.cell(row=fila_datos_sucursales, column=4, value=item.get('MONTO', 0)).number_format = '#,##0.00'
                        ws.cell(row=fila_datos_sucursales, column=4).border = self.styles.border
                        ws.cell(row=fila_datos_sucursales, column=5, value=item.get('CANTIDAD', 0)).border = self.styles.border
                        ws.cell(row=fila_datos_sucursales, column=5).alignment = self.styles.center_align
                        
                        # Columna FACTURAS
                        facturas_str = item.get('FACTURAS', '')
                        num_facturas = len(facturas_str.split(',')) if facturas_str else 0
                        ws.cell(row=fila_datos_sucursales, column=6, value=facturas_str).border = self.styles.border
                        
                        # SUGERENCIA 3: DROP SIZE (MONTO / FACTURAS)
                        drop_size = item.get('MONTO', 0) / num_facturas if num_facturas > 0 else 0
                        c_drop = ws.cell(row=fila_datos_sucursales, column=7, value=drop_size)
                        c_drop.number_format = '#,##0.00'
                        c_drop.border = self.styles.border
                        c_drop.alignment = self.styles.center_align
                        
                        # SUGERENCIA 3: FRECUENCIA (FACTURAS / MESES)
                        frecuencia = num_facturas / num_meses if num_meses > 0 else 0
                        c_freq = ws.cell(row=fila_datos_sucursales, column=8, value=frecuencia)
                        c_freq.number_format = '0.0'
                        c_freq.border = self.styles.border
                        c_freq.alignment = self.styles.center_align
                        
                        # CORRECCIÓN: PCT_INDIVIDUAL viene en porcentaje, dividir entre 100 para formato decimal de Excel
                        pct_ind = item.get('PCT_INDIVIDUAL', 0) / 100
                        ws.cell(row=fila_datos_sucursales, column=9, value=pct_ind).number_format = '0.00%'
                        ws.cell(row=fila_datos_sucursales, column=9).border = self.styles.border
                        ws.cell(row=fila_datos_sucursales, column=10, value=item.get('CATEGORIA', '')).border = self.styles.border
                        ws.cell(row=fila_datos_sucursales, column=10).alignment = self.styles.center_align
                        fila_datos_sucursales += 1
                    
                    fila_datos_lineas = fila_datos_sucursales  # Actualizar para la siguiente tabla
        
        # === TABLA D: EVOLUCIÓN MENSUAL ===
        if historial is not None:
            # Obtener periodos del historial
            if 'PERIODO_TEND' not in historial.columns:
                historial['PERIODO_TEND'] = historial['FECHA_ORIG'].dt.to_period('M').astype(str)
            
            # FIX 1: Filtrar historial por rango de fechas (consistencia con Tabla A)
            df_hist = historial.copy()
            if rango_fechas and rango_fechas[0] and rango_fechas[1]:
                df_hist = df_hist[
                    (df_hist['FECHA_ORIG'] >= rango_fechas[0]) &
                    (df_hist['FECHA_ORIG'] <= rango_fechas[1])
                ]
            
            # FIX 4: Excluir SOLES negativos (devoluciones) para consistencia con Pareto
            if 'SOLES' in df_hist.columns:
                df_hist = df_hist[df_hist['SOLES'] >= 0].copy()
            
            periodos = sorted(df_hist['PERIODO_TEND'].unique())
            
            # Filtrar historial por vendedor
            if 'ID_VENDEDOR' in df_hist.columns:
                df_vendedor = df_hist[df_hist['ID_VENDEDOR'] == id_vendedor]
            else:
                df_vendedor = df_hist
            
            if not df_vendedor.empty:
                fila_inicio_evol = fila_datos_lineas + 2
                ws.cell(row=fila_inicio_evol, column=1, value="TABLA D: EVOLUCIÓN MENSUAL").font = Font(bold=True, size=11)
                
                fila_cab_evol = fila_inicio_evol + 1
                # SUGERENCIA 4: Agregar RECENCY, FORECAST y VARIACIÓN %
                cabeceras_evol = ['N°', 'CLIENTE'] + [p.upper() for p in periodos] + ['PROM. MENSUAL', 'FORECAST', 'VAR %', 'RECENCY', 'TEND']
                for i, text in enumerate(cabeceras_evol, 1):
                    c = ws.cell(row=fila_cab_evol, column=i, value=text)
                    c.fill = self.styles.header_fill
                    c.font = self.styles.header_font
                    c.border = self.styles.border
                    c.alignment = self.styles.center_align
                
                # Agrupar por cliente y periodo
                df_mensual = df_vendedor.groupby(['ID_CLIENTE', 'NOM_CLIENTE', 'PERIODO_TEND'])['SOLES'].sum().unstack(fill_value=0)
                
                # Calcular totales por cliente para ordenar
                clientes_totales = []
                for row_key in df_mensual.index:
                    total = sum(df_mensual.loc[row_key].get(p, 0) for p in periodos)
                    clientes_totales.append((row_key, total))
                
                # Ordenar por total descendente
                clientes_totales.sort(key=lambda x: x[1], reverse=True)
                
                # Calcular fecha actual para Recency
                from datetime import datetime
                fecha_actual = datetime.now()
                
                # Escribir datos de evolución
                fila_datos_evol = fila_cab_evol + 1
                for idx, (row_key, _) in enumerate(clientes_totales):
                    id_c, nom_c = row_key
                    nro = idx + 1
                    
                    # Verificar si es cliente VITAL
                    es_vital = False
                    id_c_str = str(id_c)
                    for cli in clientes:
                        if str(cli.get('ID_CLIENTE', '')) == id_c_str:
                            es_vital = 'VITAL' in str(cli.get('CATEGORIA', '')).upper()
                            break
                    
                    ws.cell(row=fila_datos_evol, column=1, value=nro).border = self.styles.border
                    ws.cell(row=fila_datos_evol, column=1).alignment = self.styles.center_align
                    ws.cell(row=fila_datos_evol, column=2, value=format_id_name(id_c, nom_c)).border = self.styles.border
                    ws.cell(row=fila_datos_evol, column=2).alignment = self.styles.left_align
                    
                    c_vals = 3
                    sum_soles = 0
                    previous_month_value = 0
                    
                    # Promedio mensual del cliente (para umbral de significancia en huecos)
                    total_cliente_meses = df_mensual.loc[row_key].sum() if row_key in df_mensual.index else 0
                    avg_cliente_mensual = total_cliente_meses / len(periodos) if periodos else 0
                    
                    for p in periodos:
                        val = df_mensual.loc[row_key].get(p, 0)
                        sum_soles += val
                        cell = ws.cell(row=fila_datos_evol, column=c_vals, value=val)
                        cell.number_format = '#,##0.00'
                        cell.border = self.styles.border
                        cell.alignment = self.styles.right_align
                        
                        # Detección de huecos (0.00) en clientes VITALES (solo si el cliente tiene historial significativo)
                        if es_vital and val == 0 and avg_cliente_mensual >= 10:
                            cell.fill = self.styles.alert_fill
                            cell.font = self.styles.alert_font
                        
                        # Alerta de caída >20%
                        if previous_month_value > 0 and (val - previous_month_value) / previous_month_value < -0.20:
                            cell.fill = self.styles.warning_fill
                            cell.font = self.styles.alert_font
                        
                        previous_month_value = val
                        c_vals += 1
                    
                    # Promedio mensual (right)
                    prom = ws.cell(row=fila_datos_evol, column=c_vals, value=sum_soles / len(periodos) if periodos else 0)
                    prom.number_format = '#,##0.00'
                    prom.border = self.styles.border
                    prom.fill = self.styles.total_fill
                    prom.alignment = self.styles.right_align
                    
                    # SUGERENCIA 4: FORECAST (Proyección basada en PROM. MENSUAL)
                    # CORRECCIÓN: Manejar datos incompletos
                    c_vals += 1
                    if periodos:
                        # Detectar si el último periodo es el mes actual del sistema
                        from datetime import datetime
                        mes_actual_sistema = datetime.now().strftime('%Y-%m')
                        mes_ultimo_periodo = periodos[-1] if periodos else ''
                        
                        # Si el último periodo es el mes actual y el monto es muy bajo, podría estar incompleto
                        if mes_ultimo_periodo == mes_actual_sistema and sum_soles / len(periodos) < df_mensual.loc[row_key].get(periodos[-1], 0) * 0.5:
                            forecast = 0  # Datos incompletos, no mostrar proyección
                        else:
                            forecast = prom.value * 1  # Proyección simple: PROM × 1 mes
                    else:
                        forecast = 0
                    
                    c_forecast = ws.cell(row=fila_datos_evol, column=c_vals, value=forecast)
                    c_forecast.number_format = '#,##0.00'
                    c_forecast.border = self.styles.border
                    c_forecast.fill = self.styles.kpi_value_fill
                    c_forecast.alignment = self.styles.right_align
                    
                    # SUGERENCIA 4: VARIACIÓN % (Mes actual vs anterior)
                    # CORRECCIÓN: Manejar datos incompletos y división por cero
                    c_vals += 1
                    if len(periodos) >= 2:
                        mes_actual = df_mensual.loc[row_key].get(periodos[-1], 0)
                        mes_anterior = df_mensual.loc[row_key].get(periodos[-2], 0)
                        
                        # Si el mes actual es el último y está incompleto (ej. '05-MAYO'), devolver 'En proceso'
                        # Detectar si el mes actual es el mes actual del sistema
                        from datetime import datetime
                        mes_actual_sistema = datetime.now().strftime('%Y-%m')
                        mes_actual_periodo = periodos[-1] if periodos else ''
                        
                        # Si el mes actual es el mes del sistema y el monto es muy bajo, podría estar incompleto
                        if mes_actual_periodo == mes_actual_sistema and mes_actual < mes_anterior * 0.5:
                            var_pct = 0  # Datos incompletos, no mostrar variación
                        elif mes_anterior > 0:
                            var_pct = (mes_actual - mes_anterior) / mes_anterior
                        else:
                            var_pct = 0  # División por cero
                    else:
                        var_pct = 0  # No hay suficientes datos
                    
                    c_var = ws.cell(row=fila_datos_evol, column=c_vals, value=var_pct)
                    c_var.number_format = '0.00%'
                    c_var.border = self.styles.border
                    c_var.alignment = self.styles.center_align
                    if var_pct < -0.20:
                        c_var.fill = self.styles.alert_fill
                        c_var.font = self.styles.alert_font
                    elif var_pct > 0.20:
                        c_var.fill = self.styles.total_fill
                    
                    # Calcular tendencia ANTES para que esté disponible en alertas
                    if len(periodos) >= 2:
                        primer_mes = df_mensual.loc[row_key].get(periodos[0], 0)
                        ultimo_mes = df_mensual.loc[row_key].get(periodos[-1], 0)
                        if ultimo_mes > primer_mes * 1.05:
                            tend = '🔺'
                        elif ultimo_mes < primer_mes * 0.95:
                            tend = '🔻'
                        else:
                            tend = '➡️'
                    else:
                        tend = '➡️'
                    
                    # SUGERENCIA 4: RECENCY (Días desde última compra)
                    c_vals += 1
                    # Obtener fecha de última compra del historial
                    df_cliente = df_vendedor[(df_vendedor['ID_CLIENTE'] == id_c) & (df_vendedor['SOLES'] > 0)]
                    if not df_cliente.empty:
                        ultima_fecha = df_cliente['FECHA_ORIG'].max()
                        if pd.notna(ultima_fecha):
                            recency = (fecha_actual - ultima_fecha).days
                        else:
                            recency = 999
                    else:
                        recency = 999
                    
                    c_recency = ws.cell(row=fila_datos_evol, column=c_vals, value=recency)
                    c_recency.number_format = '#,##0'
                    c_recency.border = self.styles.border
                    c_recency.alignment = self.styles.center_align
                    # Alerta de Fuga de Cliente: si Recency > 30 días y es VITAL
                    if es_vital and recency > 30:
                        c_recency.fill = self.styles.alert_fill
                        c_recency.font = self.styles.alert_font
                    # Alerta CRÍTICA: VITAL + recency >60 + tendencia 🔻
                    if es_vital and recency > 60 and '🔻' in tend:
                        c_recency.fill = self.styles.hhi_high_fill
                        c_recency.font = self.styles.critical_white_font
                    
                    # Tendencia
                    c_vals += 1
                    c_tend = ws.cell(row=fila_datos_evol, column=c_vals, value=tend)
                    c_tend.border = self.styles.border
                    c_tend.alignment = self.styles.center_align
                    if '🔺' in tend:
                        c_tend.font = self.styles.trend_up_font
                    elif '🔻' in tend:
                        c_tend.font = self.styles.trend_down_font
                    # FIX 3: Marcar en rojo TEND de clientes VITALES críticos (recency alto + caída)
                    if es_vital and recency > 60 and '🔻' in tend:
                        c_tend.fill = self.styles.hhi_high_fill
                        c_tend.font = self.styles.critical_white_font
                    
                    fila_datos_evol += 1
        
        # === FILA DE DEVOLUCIONES EN TABLA D ===
        if devoluciones_por_periodo and periodos:
            fila_dev = fila_datos_evol
            dev_col_start = 3
            for dc in range(1, 3 + len(periodos) + 5):
                ws.cell(row=fila_dev, column=dc).border = self.styles.border
            ws.cell(row=fila_dev, column=2, value='DEVOLUCIONES (NC)').font = self.styles.dev_label_font
            total_dev_periodos = 0
            for pi, p in enumerate(periodos):
                val_dev = devoluciones_por_periodo.get(p, 0)
                total_dev_periodos += val_dev
                cell_dev = ws.cell(row=fila_dev, column=dev_col_start + pi, value=val_dev)
                cell_dev.number_format = '#,##0.00'
                cell_dev.font = self.styles.dev_font
            prom_dev = total_dev_periodos / len(periodos) if periodos else 0
            c_prom_dev = ws.cell(row=fila_dev, column=dev_col_start + len(periodos), value=prom_dev)
            c_prom_dev.number_format = '#,##0.00'
            c_prom_dev.font = self.styles.dev_sub_font
            fila_datos_evol = fila_dev + 1
        
        # === LEYENDA / CARTILLA DE INTERPRETACIÓN ===
        fila_leyenda = fila_datos_evol + 2 if historial is not None and not df_vendedor.empty else fila_datos_lineas + 2
        ws.cell(row=fila_leyenda, column=1, value="CÓMO LEER ESTE REPORTE").font = Font(bold=True, size=11, color="0D2B4E")
        fila_leyenda += 1
        leyendas = [
            ("ICONOS", "🔺 Crecimiento (>5%)  |  🔻 Caída (>5%)  |  ➡️ Estable  |  ⏳ Mes en proceso"),
            ("COLORES", "🔴 Rojo = Alerta / Atención  |  🟡 Ámbar = Precaución  |  🟢 Verde = Saludable"),
            ("CATEGORÍAS", "VITAL (≤80%) = Clientes que generan el 80% de la facturación  |  TRIVIAL (>80%) = Clientes menores"),
            ("VULNERABILIDAD", "ALTA = >80% de su compra en una sola línea  |  MEDIA = 60-80%  |  BAJA = <60% (diversificado)"),
            ("HHI (KPI)", "<0.15 = Baja concentración (saludable)  |  0.15-0.25 = Moderada  |  >0.25 = ALTA (riesgo de dependencia)"),
            ("DEV%", "Devoluciones / Venta Total. >5% = Alerta (impacta comisión neta)"),
            ("FÓRMULAS", "Las celdas de % contienen fórmulas de Excel. Use la fila TOTAL (SUBTOTAL) para filtrar por categoría."),
        ]
        for label, text in leyendas:
            ws.cell(row=fila_leyenda, column=1, value=label).font = Font(bold=True, size=9, color="0D2B4E")
            ws.cell(row=fila_leyenda, column=1).border = self.styles.border
            ws.cell(row=fila_leyenda, column=2, value=text).font = Font(size=9)
            ws.cell(row=fila_leyenda, column=2).border = self.styles.border
            ws.merge_cells(start_row=fila_leyenda, start_column=2, end_row=fila_leyenda, end_column=8)
            fila_leyenda += 1
        
        # Auto-ajuste de columnas
        self._auto_adjust_columns(ws)
        ws.freeze_panes = "B6"

    def _escribir_resumen_consolidado(self, ws, titulo, datos_por_vendedor, rango_fechas, vendedores_nombres=None, filtros_aplicados: str = ""):
        ws.cell(row=1, column=1, value=titulo).font = self.styles.title_font
        ws.cell(row=2, column=1, value=f"Fecha: {datetime.now().strftime('%d/%m/%Y')}")
        if rango_fechas and rango_fechas[0] and rango_fechas[1]:
            ws.cell(row=3, column=1, value=f"Rango: {rango_fechas[0].strftime('%d/%m/%Y')} -> {rango_fechas[1].strftime('%d/%m/%Y')}")
        
        if filtros_aplicados:
            ws.cell(row=4, column=1, value=f"Filtros: {filtros_aplicados}").font = self.styles.italic_gray_font

        if vendedores_nombres is None:
            vendedores_nombres = {}
        cabeceras = ['N°', 'VENDEDOR', 'ITEMS', 'MONTO_TOTAL']
        for i, cab in enumerate(cabeceras, 1):
            celda = ws.cell(row=5, column=i, value=cab)
            celda.fill = self.styles.header_fill
            celda.font = self.styles.header_font
            celda.alignment = self.styles.center_align
            celda.border = self.styles.border
        fila = 6
        for idx, (vendedor_id, datos) in enumerate(datos_por_vendedor.items()):
            total = sum(d.get('MONTO', 0) for d in datos)
            nombre = vendedores_nombres.get(vendedor_id, "")
            nro = idx + 1
            ws.cell(row=fila, column=1, value=nro).border = self.styles.border
            ws.cell(row=fila, column=1).alignment = self.styles.center_align
            ws.cell(row=fila, column=2, value=format_id_name(vendedor_id, nombre, field_name='VENDEDOR')).border = self.styles.border
            ws.cell(row=fila, column=3, value=len(datos)).border = self.styles.border
            ws.cell(row=fila, column=3).alignment = self.styles.center_align
            ws.cell(row=fila, column=4, value=total).number_format = '#,##0.00'
            ws.cell(row=fila, column=4).border = self.styles.border
            fila += 1

    def _escribir_encabezado_consolidado(self, ws, titulo, vendedor_id, rango_fechas, vendedor_nombre=""):
        ws.cell(row=1, column=1, value=titulo).font = self.styles.title_font
        ws.cell(row=2, column=1, value=f"Vendedor: {format_id_name(vendedor_id, vendedor_nombre, field_name='VENDEDOR')}")
        ws.cell(row=3, column=1, value=f"Fecha: {datetime.now().strftime('%d/%m/%Y')}")
        if rango_fechas and rango_fechas[0] and rango_fechas[1]:
            ws.cell(row=4, column=1, value=f"Periodo: {rango_fechas[0].strftime('%d/%m/%Y')} -> {rango_fechas[1].strftime('%d/%m/%Y')}")

    def _detectar_cabeceras(self, dato_ejemplo: dict, agrupacion: str) -> list:
        """Detecta los campos disponibles y genera encabezados dinámicamente según el tipo de agrupación."""
        campos_presentes = set(dato_ejemplo.keys())
        
        # Orden de columnas según tipo de agrupación
        if agrupacion == "ID_CLIENTE":
            # Por CLIENTE: CLIENTE -> LÍNEA -> SKU -> CANTIDAD -> MONTO -> FACTURAS -> PRECIOS -> PEDIDOS -> FECHAS
            orden = [
                'CLIENTE', 'LÍNEA', 'SKU',
                'CANTIDAD', 'MONTO',
                'FACTURAS', 'PRECIOS', 'PEDIDOS',
                'FECHA_ULT', 'FECHA_MIN'
            ]
        elif agrupacion == "NOM_LINEA":
            # Por LÍNEA: LÍNEA -> SKU -> CLIENTE -> CANTIDAD -> MONTO -> FACTURAS -> PRECIOS -> PEDIDOS -> FECHAS
            orden = [
                'LÍNEA', 'SKU', 'CLIENTE',
                'CANTIDAD', 'MONTO',
                'FACTURAS', 'PRECIOS', 'PEDIDOS',
                'FECHA_ULT', 'FECHA_MIN'
            ]
        elif agrupacion == "ID_ARTICULO":
            # Por SKU: SKU -> LÍNEA -> CLIENTE -> CANTIDAD -> MONTO -> FACTURAS -> PRECIOS -> PEDIDOS -> FECHAS
            orden = [
                'SKU', 'LÍNEA', 'CLIENTE',
                'CANTIDAD', 'MONTO',
                'FACTURAS', 'PRECIOS', 'PEDIDOS',
                'FECHA_ULT', 'FECHA_MIN'
            ]
        elif agrupacion == "PERIODO_MES":
            # Por PERIODO: PERIODO -> TIPO_PERIODO -> SKU -> LÍNEA -> CLIENTE -> CANTIDAD -> MONTO -> FACTURAS -> PRECIOS -> PEDIDOS -> FECHAS
            orden = [
                'PERIODO', 'TIPO_PERIODO', 'SKU', 'LÍNEA', 'CLIENTE',
                'CANTIDAD', 'MONTO',
                'FACTURAS', 'PRECIOS', 'PEDIDOS',
                'FECHA_ULT', 'FECHA_MIN', 'FECHA_MAX'
            ]
        elif agrupacion == "FACTURA":
            # Por FACTURA: FACTURA -> FECHA -> CLIENTE -> LÍNEA -> SKU -> CANTIDAD -> PRECIO -> MONTO -> PEDIDO
            orden = [
                'FACTURA', 'FECHA', 'CLIENTE', 'LÍNEA', 'SKU',
                'CANTIDAD', 'PRECIO', 'MONTO', 'PEDIDO'
            ]
        else:
            # Por defecto: SKU -> LÍNEA -> CLIENTE -> CANTIDAD -> MONTO -> FACTURAS -> PRECIOS -> PEDIDOS -> FECHAS
            orden = [
                'SKU', 'LÍNEA', 'CLIENTE',
                'CANTIDAD', 'MONTO',
                'FACTURAS', 'PRECIOS', 'PEDIDOS',
                'FECHA_ULT', 'FECHA_MIN'
            ]
        
        cabeceras = []
        
        # Agregar campos en orden de prioridad si existen
        for campo in orden:
            if campo in campos_presentes:
                cabeceras.append(campo)
        
        # Si no se detectó ningún campo, usar encabezados por defecto
        if not cabeceras:
            cabeceras = ['CLIENTE', 'LÍNEA', 'SKU', 'CANTIDAD', 'MONTO']
        
        return cabeceras

    def _escribir_fila_consolidado(self, ws, fila: int, dato: dict, cabeceras: list):
        """Escribe una fila de datos según los encabezados detectados."""
        for col_idx, encabezado in enumerate(cabeceras, 1):
            valor = dato.get(encabezado, '')
            
            # Aplicar formato numérico según el tipo de campo
            if encabezado in ['CANTIDAD', 'CANT'] or (encabezado and '-CANT' in encabezado):
                cell = ws.cell(row=fila, column=col_idx, value=valor)
                cell.number_format = '#,##0'
            elif encabezado in ['MONTO', 'PRECIO', 'P.U.'] or (encabezado and '-MONTO' in encabezado) or (encabezado and '-PRECIO' in encabezado):
                cell = ws.cell(row=fila, column=col_idx, value=valor)
                cell.number_format = '#,##0.00'
            elif 'DIF' in encabezado.upper():
                cell = ws.cell(row=fila, column=col_idx, value=valor)
                cell.number_format = '#,##0.00'
            elif '%' in encabezado or 'PCT' in encabezado.upper() or 'ACUM' in encabezado.upper():
                if isinstance(valor, (int, float)):
                    valor = valor / 100 if valor > 1 else valor
                cell = ws.cell(row=fila, column=col_idx, value=valor)
                cell.number_format = '0.00%'
            else:
                ws.cell(row=fila, column=col_idx, value=valor)
            
            # Aplicar borde
            ws.cell(row=fila, column=col_idx).border = self.styles.border

    def _aplicar_formato_condicional_monto(self, ws, cabeceras, max_row):
        """Aplica barras de datos a las columnas de MONTO."""
        from openpyxl.utils import get_column_letter
        for i, cab in enumerate(cabeceras, 1):
            if 'MONTO' in cab.upper():
                col_letter = get_column_letter(i)
                rule = DataBarRule(start_type='min', end_type='max', color="00D084", showValue=True)
                ws.conditional_formatting.add(f"{col_letter}6:{col_letter}{max_row}", rule)

    def _auto_adjust_columns(self, ws):
        from openpyxl.utils import get_column_letter
        for col in ws.columns:
            max_length = 0
            column_idx = col[0].column
            column = get_column_letter(column_idx)
            for i, cell in enumerate(col):
                if i > 100: break
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            ws.column_dimensions[column].width = min(max_length + 3, 50)
