import sys
import re

with open('src/excel/generator.py', 'r', encoding='utf-8') as f:
    content = f.read()

# PART 1: Fix chart categories to use both columns (CLIENTE and VENDEDOR)
old_chart = """            # Tomar los top 10 clientes de la tabla que acabamos de escribir
            num_clients = min(len(clientes), 10)
            if num_clients > 0:
                # data incluye el encabezado "TOTAL SOLES" en col 3
                data = Reference(ws, min_col=3, min_row=fila_h+1, max_row=fila_h+1+num_clients)
                # categories son los nombres de cliente col 1
                cats = Reference(ws, min_col=1, min_row=fila_h+2, max_row=fila_h+1+num_clients)
                
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(cats)
                chart.shape = 4"""

new_chart = """            # Tomar los top 10 clientes de la tabla que acabamos de escribir
            num_clients = min(len(clientes), 10)
            if num_clients > 0:
                # data incluye el encabezado "TOTAL SOLES" en col 3
                data = Reference(ws, min_col=3, min_row=fila_h+1, max_row=fila_h+1+num_clients)
                # categories son los nombres de cliente y vendedor (col 1 y 2) para agrupación
                cats = Reference(ws, min_col=1, min_row=fila_h+2, max_col=2, max_row=fila_h+1+num_clients)
                
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(cats)
                chart.shape = 4"""

content = content.replace(old_chart, new_chart)


# PART 2: Refactor _escribir_anexo_evolucion to create one sheet per vendor

old_evolucion = """    def _escribir_anexo_evolucion(self, datos_pareto, historial, rango_fechas):
        \"\"\"Hoja para medir los meses presentes en el historial (Análisis de Evolución).\"\"\"
        ws = self.wb.create_sheet("EVOLUCION_MENSUAL")
        periodos = datos_pareto.get('PERIODOS', [])
        
        ws.cell(row=1, column=1, value="ANEXO: EVOLUCIÓN MENSUAL DE CLIENTES").font = self.styles.title_font
        
        cabeceras = ['CLIENTE', 'VENDEDOR'] + [p.upper() for p in periodos] + ['PROM. MENSUAL']
        for i, text in enumerate(cabeceras, 1):
            c = ws.cell(row=4, column=i, value=text)
            c.fill = self.styles.header_fill; c.font = self.styles.header_font; c.border = self.styles.border

        # Agrupar historial por Cliente y Mes
        # Asegurar que la columna de periodo exista para el agrupamiento
        if 'PERIODO_TEND' not in historial.columns:
            historial['PERIODO_TEND'] = historial['FECHA_ORIG'].dt.to_period('M').astype(str)
        
        # Verificar si existen las columnas de vendedor
        if 'ID_VENDEDOR' in historial.columns and 'NOM_VENDEDOR' in historial.columns:
            df_mensual = historial.groupby(['ID_CLIENTE', 'NOM_CLIENTE', 'ID_VENDEDOR', 'NOM_VENDEDOR', 'PERIODO_TEND'])['SOLES'].sum().unstack(fill_value=0)
            usar_id_vendedor = True
        else:
            df_mensual = historial.groupby(['ID_CLIENTE', 'NOM_CLIENTE', 'NOM_VENDEDOR', 'PERIODO_TEND'])['SOLES'].sum().unstack(fill_value=0)
            usar_id_vendedor = False
        
        fila = 5
        for row_key, row_vals in df_mensual.iterrows():
            if usar_id_vendedor:
                id_c, nom_c, id_v, nom_v = row_key
                ws.cell(row=fila, column=2, value=format_id_name(id_v, nom_v)).border = self.styles.border
            else:
                id_c, nom_c, nom_v = row_key
                ws.cell(row=fila, column=2, value=str(nom_v) if nom_v else "").border = self.styles.border
            
            ws.cell(row=fila, column=1, value=format_id_name(id_c, nom_c)).border = self.styles.border
            
            c_vals = 3
            sum_soles = 0
            previous_month_value = 0 # Inicializar para el primer mes
            
            for p in periodos:
                val = row_vals.get(p, 0)
                sum_soles += val
                cell = ws.cell(row=fila, column=c_vals, value=val)
                cell.number_format = '#,##0.00'; cell.border = self.styles.border
                
                # Aplicar formato condicional si la caída es > 20% respecto al mes anterior
                if previous_month_value > 0 and (val - previous_month_value) / previous_month_value < -0.20:
                    cell.fill = self.styles.warning_fill
                    cell.font = self.styles.alert_font
                
                previous_month_value = val # Actualizar para la siguiente iteración
                c_vals += 1
            
            prom = ws.cell(row=fila, column=c_vals, value=sum_soles / len(periodos) if periodos else 0)
            prom.number_format = '#,##0.00'; prom.border = self.styles.border; prom.fill = self.styles.total_fill
            fila += 1
            
        self._auto_adjust_columns(ws)
        ws.freeze_panes = "C5\"\"\"

new_evolucion = """    def _escribir_anexo_evolucion(self, datos_pareto, historial, rango_fechas):
        \"\"\"Hojas para medir los meses presentes en el historial (Análisis de Evolución por Vendedor).\"\"\"
        periodos = datos_pareto.get('PERIODOS', [])
        
        # Agrupar historial por Cliente y Mes
        if 'PERIODO_TEND' not in historial.columns:
            historial['PERIODO_TEND'] = historial['FECHA_ORIG'].dt.to_period('M').astype(str)
            
        if 'ID_VENDEDOR' in historial.columns and 'NOM_VENDEDOR' in historial.columns:
            df_mensual = historial.groupby(['ID_VENDEDOR', 'NOM_VENDEDOR', 'ID_CLIENTE', 'NOM_CLIENTE', 'PERIODO_TEND'])['SOLES'].sum().unstack(fill_value=0)
            usar_id_vendedor = True
        else:
            df_mensual = historial.groupby(['NOM_VENDEDOR', 'ID_CLIENTE', 'NOM_CLIENTE', 'PERIODO_TEND'])['SOLES'].sum().unstack(fill_value=0)
            usar_id_vendedor = False
            
        # Obtener los vendedores únicos
        if usar_id_vendedor:
            vendedores_keys = df_mensual.index.droplevel(['ID_CLIENTE', 'NOM_CLIENTE']).unique()
        else:
            vendedores_keys = df_mensual.index.droplevel(['ID_CLIENTE', 'NOM_CLIENTE']).unique()
            
        # Eliminar la hoja "EVOLUCION_MENSUAL" si existiera previamente
        if "EVOLUCION_MENSUAL" in self.wb.sheetnames:
            del self.wb["EVOLUCION_MENSUAL"]
            
        for v_key in vendedores_keys:
            if usar_id_vendedor:
                id_v, nom_v = v_key
                vendedor_display = format_id_name(id_v, nom_v)
                df_vendedor = df_mensual.loc[(id_v, nom_v)]
            else:
                nom_v = v_key
                vendedor_display = str(nom_v)
                df_vendedor = df_mensual.loc[nom_v]
                
            # Crear nombre de hoja seguro
            sheet_name = f"EVOL_{vendedor_display[:26]}"[:31]
            sheet_name = re.sub(r'[\\\\/*?:\\[\\]]', "", sheet_name)
            
            # Evitar duplicados
            if sheet_name in self.wb.sheetnames:
                sheet_name = f"{sheet_name}_c"[:31]
                
            ws = self.wb.create_sheet(sheet_name)
            ws.cell(row=1, column=1, value=f"EVOLUCIÓN MENSUAL: {vendedor_display}").font = self.styles.title_font
            
            cabeceras = ['CLIENTE'] + [p.upper() for p in periodos] + ['PROM. MENSUAL']
            for i, text in enumerate(cabeceras, 1):
                c = ws.cell(row=4, column=i, value=text)
                c.fill = self.styles.header_fill; c.font = self.styles.header_font; c.border = self.styles.border
                
            fila = 5
            for row_key, row_vals in df_vendedor.iterrows():
                id_c, nom_c = row_key
                ws.cell(row=fila, column=1, value=format_id_name(id_c, nom_c)).border = self.styles.border
                
                c_vals = 2
                sum_soles = 0
                previous_month_value = 0
                
                for p in periodos:
                    val = row_vals.get(p, 0)
                    sum_soles += val
                    cell = ws.cell(row=fila, column=c_vals, value=val)
                    cell.number_format = '#,##0.00'; cell.border = self.styles.border
                    
                    if previous_month_value > 0 and (val - previous_month_value) / previous_month_value < -0.20:
                        from openpyxl.styles import PatternFill, Font
                        # Use warning format
                        cell.fill = self.styles.warning_fill
                        cell.font = self.styles.alert_font
                        
                    previous_month_value = val
                    c_vals += 1
                
                prom = ws.cell(row=fila, column=c_vals, value=sum_soles / len(periodos) if periodos else 0)
                prom.number_format = '#,##0.00'; prom.border = self.styles.border; prom.fill = self.styles.total_fill
                fila += 1
                
            self._auto_adjust_columns(ws)
            ws.freeze_panes = "B5\"\"\"

content_norm = content.replace('\\r\\n', '\\n')
old_evolucion_norm = old_evolucion.replace('\\r\\n', '\\n')

if old_evolucion_norm in content_norm:
    print('Replacing anexo evolucion...')
    content_norm = content_norm.replace(old_evolucion_norm, new_evolucion.replace('\\r\\n', '\\n'))
else:
    print('Failed to replace anexo evolucion!')

with open('src/excel/generator.py', 'w', encoding='utf-8') as f:
    f.write(content_norm)
print('Done.')
